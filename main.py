import os
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from fastapi import FastAPI, Form, UploadFile, File, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from io import BytesIO
import logging
from collections import defaultdict
import re

app = FastAPI()
logger = logging.getLogger(__name__)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://main.d32bb122o9jw4d.amplifyapp.com"],
    allow_methods=["*"],
    allow_headers=["*"],
)

def generate_resumen_tecnicos(writer, datos_combinados, df_originales):

    instalados_por_tecnico = {}
    retirados_por_tecnico = {}
    luminarias_por_tecnico = {}
    tecnicos = set()
    
    # Mapeo de nodos a técnicos
    nodo_tecnico_map = {}
    
    # Primero, construir un mapa de nodo -> técnico desde los DataFrames originales
    for df in df_originales.values():
        if "4.Nombre del Técnico Instalador" in df.columns and "1.NODO DEL POSTE." in df.columns:
            for _, fila in df.iterrows():
                nodo = str(fila["1.NODO DEL POSTE."]).strip()
                tecnico = str(fila.get("4.Nombre del Técnico Instalador", "Sin técnico")).strip()
                
                # Normalizar el nombre del técnico (eliminar espacios extra, capitalizar correctamente)
                tecnico = " ".join([palabra.capitalize() for palabra in tecnico.split() if palabra])
                
                if not tecnico or tecnico.lower() in ['na', 'n/a', 'ninguno']:
                    tecnico = "Sin técnico"
                
                # Manejo de nodos 0
                ot = fila["2.Nro de O.T."]
                if nodo in ['0', '0.0']:
                    nodo = f"0_{ot}"  # Usar el formato que se usa en procesar_archivo_modernizacion
                
                nodo_tecnico_map[nodo] = tecnico
                tecnicos.add(tecnico)
    
    # Ordenar técnicos para mantener consistencia
    tecnicos_ordenados = sorted(tecnicos)
    
    # Ahora procesar datos combinados
    for ot, info in datos_combinados.items():
        # Proceso para códigos N1
        for key, nodos_data in info.get('codigos_n1', {}).items():
            nombre = f"LUMINARIA {key}"
            
            for nodo, codigos in nodos_data.items():
                tecnico = nodo_tecnico_map.get(nodo.split('_')[0], "Sin técnico")
                
                if tecnico not in luminarias_por_tecnico:
                    luminarias_por_tecnico[tecnico] = {}
                
                if nombre not in luminarias_por_tecnico[tecnico]:
                    luminarias_por_tecnico[tecnico][nombre] = {
                        'tipo': 'LUMINARIA',
                        'unidad': 'UND',
                        'total': 0,
                        'ots': defaultdict(int)
                    }
                
                # Contar el número de códigos en este nodo
                codigos_count = len(codigos)
                luminarias_por_tecnico[tecnico][nombre]['total'] += codigos_count
                luminarias_por_tecnico[tecnico][nombre]['ots'][ot] += codigos_count
        
        # Proceso para códigos N2
        for key, nodos_data in info.get('codigos_n2', {}).items():
            nombre = f"LUMINARIA {key}"
            
            for nodo, codigos in nodos_data.items():
                tecnico = nodo_tecnico_map.get(nodo.split('_')[0], "Sin técnico")
                
                if tecnico not in luminarias_por_tecnico:
                    luminarias_por_tecnico[tecnico] = {}
                
                if nombre not in luminarias_por_tecnico[tecnico]:
                    luminarias_por_tecnico[tecnico][nombre] = {
                        'tipo': 'LUMINARIA',
                        'unidad': 'UND',
                        'total': 0,
                        'ots': defaultdict(int)
                    }
                
                # Contar el número de códigos en este nodo
                codigos_count = len(codigos)
                luminarias_por_tecnico[tecnico][nombre]['total'] += codigos_count
                luminarias_por_tecnico[tecnico][nombre]['ots'][ot] += codigos_count
            
        # Procesar materiales instalados
        for mat_key, nodo_quantities in info['materiales'].items():
            _, nombre = mat_key.split('|', 1)
            
            if nombre.strip().upper() == "NINGUNO":
                continue
            
            for nodo, cantidad in nodo_quantities.items():
                tecnico = nodo_tecnico_map.get(nodo.split('_')[0], "Sin técnico")
                
                if tecnico not in instalados_por_tecnico:
                    instalados_por_tecnico[tecnico] = {}
                
                if nombre not in instalados_por_tecnico[tecnico]:
                    instalados_por_tecnico[tecnico][nombre] = {
                        'tipo': 'INSTALADO',
                        'unidad': 'UND',
                        'total': 0,
                        'ots': defaultdict(int)
                    }
                instalados_por_tecnico[tecnico][nombre]['total'] += cantidad
                instalados_por_tecnico[tecnico][nombre]['ots'][ot] += cantidad
        
        # Procesar materiales retirados
        for mat_key, nodo_quantities in info.get('materiales_retirados', {}).items():
            _, nombre = mat_key.split('|', 1)
            
            if nombre.strip().upper() == "NINGUNO":
                continue
            
            for nodo, cantidad in nodo_quantities.items():
                tecnico = nodo_tecnico_map.get(nodo.split('_')[0], "Sin técnico")
                
                if tecnico not in retirados_por_tecnico:
                    retirados_por_tecnico[tecnico] = {}
                
                if nombre not in retirados_por_tecnico[tecnico]:
                    retirados_por_tecnico[tecnico][nombre] = {
                        'tipo': 'RETIRADO',
                        'unidad': 'UND',
                        'total': 0,
                        'ots': defaultdict(int)
                    }
                retirados_por_tecnico[tecnico][nombre]['total'] += cantidad
                retirados_por_tecnico[tecnico][nombre]['ots'][ot] += cantidad

    # Verificación de totales
    for tecnico, materiales in luminarias_por_tecnico.items():
        for nombre, info in materiales.items():
            # Recalcular el total sumando todas las cantidades por OT para confirmar
            total_from_ots = sum(info['ots'].values())
            if info['total'] != total_from_ots:
                # Ajustar si hay discrepancia
                info['total'] = total_from_ots
    
    # Crear la hoja para el resumen de técnicos
    sheet_name = 'Resumen_tecnicos'
    worksheet = writer.book.create_sheet(sheet_name)
    
    # Configuración de estilos
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Columnas del resumen
    columns = ['Nombre Material', 'Unidad', 'Cantidad Total'] + tecnicos_ordenados
    
    # Si no hay datos, salir
    if not (luminarias_por_tecnico or instalados_por_tecnico or retirados_por_tecnico):
        return
    
    # Escribir encabezados de columnas
    for col, header in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    current_row = 2
    
    # Función para escribir una sección
    def write_section(title, data_by_tecnico, row):
        # Encabezado de sección
        cell = worksheet.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_alignment
        
        # Combinar celdas para el encabezado
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        
        row += 1
        
        # Recolectar todos los materiales únicos para esta sección
        all_materials = set()
        for tecnico in tecnicos_ordenados:
            if tecnico in data_by_tecnico:
                all_materials.update(data_by_tecnico[tecnico].keys())
        
        # Escribir datos por material
        for material in sorted(all_materials):
            worksheet.cell(row=row, column=1, value=material)
            worksheet.cell(row=row, column=2, value='UND')
            
            # Calcular el total sumando de todos los técnicos
            total = sum(
                data_by_tecnico.get(tecnico, {}).get(material, {}).get('total', 0)
                for tecnico in tecnicos_ordenados
            )
            worksheet.cell(row=row, column=3, value=total)
            
            # Cantidades por técnico
            for i, tecnico in enumerate(tecnicos_ordenados, 4):
                cantidad = data_by_tecnico.get(tecnico, {}).get(material, {}).get('total', 0)
                worksheet.cell(row=row, column=i, value=cantidad)
            
            row += 1
            
        return row

    # Escribir sección de luminarias
    if luminarias_por_tecnico:
        current_row = write_section("LUMINARIAS INSTALADAS", luminarias_por_tecnico, current_row)
        current_row += 1  # Espacio entre secciones
    
    # Escribir sección de materiales instalados
    if instalados_por_tecnico:
        current_row = write_section("MATERIALES INSTALADOS", instalados_por_tecnico, current_row)
        current_row += 1  # Espacio entre secciones
    
    # Escribir sección de materiales retirados
    if retirados_por_tecnico:
        current_row = write_section("MATERIALES RETIRADOS", retirados_por_tecnico, current_row)
    
    # Ajustar anchos de columna
    for idx, col in enumerate(columns, 1):
        col_width = max(len(str(col)), 15) + 2
        # Para la primera columna, usar un ancho mayor
        if idx == 1:
            col_width = 40
        worksheet.column_dimensions[get_column_letter(idx)].width = col_width
    
    # Configurar congelación de paneles
    worksheet.freeze_panes = 'D2'

def extraer_cantidad(texto):
    """
    Extrae la cantidad numérica de un texto.
    
    Formatos soportados:
    - 'Descripción (cantidad)'
    - 'Descripción cantidad UND'
    - 'cantidad UND'
    
    Args:
        texto: Texto del que extraer la cantidad
        
    Returns:
        float: Cantidad extraída o 0 si no se encuentra
    """
    try:
        # Caso 1: Formato 'Descripción (cantidad)'
        if '(' in texto and ')' in texto:
            cantidad_str = texto.split('(')[-1].rstrip(')')
            return float(cantidad_str)
        
        # Caso 2: Buscar patrón de número seguido de UND
        palabras = texto.split()
        for i, palabra in enumerate(palabras):
            if palabra.upper() in ["UND", "UN", "ML", "M", "KG", "MT", "MTS"]:
                if i > 0 and palabras[i-1].replace(',', '.').replace('-', '').isdigit():
                    return float(palabras[i-1].replace(',', '.'))
                elif i > 0:
                    try:
                        return float(palabras[i-1].replace(',', '.'))
                    except ValueError:
                        pass
        
        # Caso 3: Intentar encontrar cualquier número en el texto
        import re
        numeros = re.findall(r'\d+(?:\.\d+)?', texto.replace(',', '.'))
        if numeros:
            return float(numeros[0])
            
        return 0
    except ValueError:
        return 0  # En caso de error, devolver 0

def normalizar_barrio(barrio):
    barrio_str = str(barrio).strip()
    try:
        pd.to_datetime(barrio_str, dayfirts=True)
        return 'Sin barrio'
    except Exception:
        pass
    if barrio_str in ['0', '0.0']:
        return 'Sin barrio'
    return barrio_str.title().strip()

           
def procesar_archivo_modernizacion(file: UploadFile):
    try:
        contenido = file.file.read()
        xls = pd.ExcelFile(BytesIO(contenido))
        
        datos = defaultdict(lambda: {
            'nodos': [],
            'nodo_counts': defaultdict(int),
            'codigos_n1': defaultdict(lambda: defaultdict(set)),
            'codigos_n2': defaultdict(lambda: defaultdict(set)),
            'materiales': defaultdict(lambda: defaultdict(int)),    
            'materiales_retirados': defaultdict(lambda: defaultdict(int)),
            'aspectos_materiales': defaultdict(lambda: defaultdict(lambda: set())),
            'aspectos_retirados': defaultdict(lambda: defaultdict(lambda: set())),
            'fechas_sync': defaultdict(str),  # Fechas de sincronización
            'tipos_suelo': defaultdict(str)   # Nuevo campo para almacenar tipos de suelo
        })                 
        
        datos_por_barrio = defaultdict(lambda: {
            'materiales_instalados': defaultdict(lambda: defaultdict(int)),
            'materiales_retirados': defaultdict(lambda: defaultdict(int))
                                                })
        dfs_originales = {}
        counter_0 = defaultdict(int)
        # Configurar columnas         
        COL_INICIO = "BH"
        COL_FIN = "BO"
        idx_inicio = column_index_from_string(COL_INICIO) - 1
        idx_fin = column_index_from_string(COL_FIN) - 1
        
        # Diccionario para normalizar nodos
        nodos_normalizados = {}
        
        # Diccionarios para rastrear nodos con códigos y brazos
        nodos_con_codigos = defaultdict(lambda: defaultdict(dict))  # Estructura: ot -> nodo -> {'n1': {'codigo': X, 'potencia': Y}, 'n2': {...}}
        nodos_con_brazos = defaultdict(dict)
        
        for hoja in xls.sheet_names:
            df = xls.parse(hoja)
            
            # Validar columnas requeridas
            required_columns = {
                "2.Nro de O.T.", "1.NODO DEL POSTE.",
                "2.CODIGO DE LUMINARIA INSTALADA N1.", "3.POTENCIA DE LUMINARIA INSTALADA (W)",
                "6.CODIGO DE LUMINARIA INSTALADA N2.", "7.POTENCIA DE LUMINARIA INSTALADA (W)",
                "1. Describa Aspectos que Considere se deben tener en cuenta.",
                "FechaSincronizacion"
            }
            if not required_columns.issubset(df.columns):
                continue          
            
            dfs_originales[hoja] = df
            # Parsear y ordenar por FechaSincronizacion
            df['FechaSincronizacion'] = (
                df['FechaSincronizacion']
                .astype(str)
                .str.replace('a. m.', 'AM', case=False)
                .str.replace('p. m.', 'PM', case=False)
            )
            df['FechaSincronizacion'] = pd.to_datetime(
                df['FechaSincronizacion'],
                format='%d/%m/%Y %I:%M:%S %p',
                errors='coerce'
            )
            df = df.sort_values(by='FechaSincronizacion', ascending=True)
            # PROCESAR MATERIALES RETIRADOS
            pattern_codigo = re.compile(r'^\d+\.CODIGO DE (LUMINARIA|BOMBILLA|FOTOCELDA) RETIRADA (N\d+)\.?$', re.IGNORECASE)
            pattern_potencia = re.compile(r'^\d+\.POTENCIA DE (LUMINARIA|BOMBILLA) RETIRADA (N\d+)\.?\(W\)$', re.IGNORECASE)
            
            codigo_columns = {}
            potencia_columns = {}
            
            for col in df.columns:
                col_str = str(col).strip()
                # PORECESAR CODIGOS RETIRADA
                codigo_match = pattern_codigo.match(col_str)
                if codigo_match:
                    tipo = codigo_match.group(1).upper()
                    n = codigo_match.group(2).upper()
                    codigo_columns[(tipo, n)] = col_str
                else:
                    # PROCESAR COLUMNAS POTENCIA
                    potencia_match = pattern_potencia.match(col_str)
                    if potencia_match:
                        tipo = potencia_match.group(1).upper()
                        n = potencia_match.group(2).upper()
                        potencia_columns[(tipo, n)] = col_str
            
            # ========== PROCESAR MATERIALES ORIGINALES (MATERIAL X - CANTIDAD X) ==========
            material_cols = [col for col in df.columns if re.match(r'^(MATERIAL|Material)\s\d+$', col)]
            cantidad_cols = [col for col in df.columns if re.match(r'^(CANTIDAD MATERIAL|CANTIDAD DE MATERIAL)\s\d+$', col)]
            
            columnas_bh_bo = []
            if len(df.columns) > idx_fin:
                columnas_bh_bo = df.columns[idx_inicio : idx_fin + 1]
                        
            for _, fila in df.iterrows():
                ot = fila["2.Nro de O.T."]   
                
                barrio = fila.get("3.Barrio", "")
                barrio_normalizado = normalizar_barrio(barrio)
                                     
                original_nodo = str(fila["1.NODO DEL POSTE."]).strip().replace(' ', '').replace('-', '').upper()
                
                # Guardar la fecha de sincronización para este nodo/registro
                fecha_sincronizacion = fila["FechaSincronizacion"]
    
                # Normalizar nodos con valor 0 o similares
                if original_nodo in ['0', '0.0', '0.00', 'nan', 'NaN', 'None', '', 'NO', 'NO APLICA']:
                    # Usar un formato consistente para todos los nodos "0"
                    counter_0[ot] += 1
                    nodo = f"0_{ot}_{counter_0[ot]}"  # Formato: 0_OT_contador
                else:
                    # Normalizar el formato del nodo para evitar duplicados por formato
                    nodo_normalizado = original_nodo.replace('.0', '').replace('.00', '')
                    
                    # Verificar si este nodo ya ha sido normalizado antes
                    if (ot, nodo_normalizado) in nodos_normalizados:
                        nodo = nodos_normalizados[(ot, nodo_normalizado)]
                    else:
                        count = datos[ot]['nodo_counts'][nodo_normalizado] + 1
                        datos[ot]['nodo_counts'][nodo_normalizado] = count
                        nodo = f"{nodo_normalizado}_{count}" if count > 1 else nodo_normalizado
                        nodos_normalizados[(ot, nodo_normalizado)] = nodo
                
                datos[ot]['nodos'].append(nodo)
                
                # Guardar la fecha de sincronización para este nodo
                if pd.notna(fecha_sincronizacion):
                    datos[ot]['fechas_sync'][nodo] = fecha_sincronizacion.strftime('%d/%m/%Y %H:%M:%S')
                else:
                    datos[ot]['fechas_sync'][nodo] = "Sin fecha"
                
                # Capturar el tipo de suelo si existe la columna
                if "2.Tipo de Suelo" in df.columns:
                    tipo_suelo = fila.get("2.Tipo de Suelo", "")
                    if pd.notna(tipo_suelo):
                        datos[ot]['tipos_suelo'][nodo] = str(tipo_suelo).strip()
                                    
                codigo_n1 = fila["2.CODIGO DE LUMINARIA INSTALADA N1."]
                potencia_n1 = fila["3.POTENCIA DE LUMINARIA INSTALADA (W)"]
                
                codigo_n2 = fila["6.CODIGO DE LUMINARIA INSTALADA N2."]
                potencia_n2 = fila["7.POTENCIA DE LUMINARIA INSTALADA (W)"]
                
                for col in columnas_bh_bo:
                    cantidad = fila[col]
                    # Limpiar NaN y convertir a 0
                    if pd.isna(cantidad):
                        cantidad = 0
                    else:
                        cantidad = int(cantidad)  # Asegurar entero
                    #if pd.notna(cantidad) and float(cantidad) > 0:
                    if cantidad > 0:
                        nombre_material = str(col).split('.', 1)[-1].strip().upper()
                        key = f"MATERIAL_RETIRADO|{nombre_material}"
                        datos[ot]['materiales_retirados'][key][nodo] += cantidad
                
                        # Capture aspect here
                        aspecto = fila["1. Describa Aspectos que Considere se deben tener en cuenta."]
                        if pd.notna(aspecto):
                            aspecto_limpio = str(aspecto).strip().upper()
                            if aspecto_limpio not in ['', 'NA', 'NINGUNO', 'N/A', 'NO', 'NO APLICA']:
                                datos[ot]['aspectos_retirados'][key][nodo].add(aspecto_limpio)
                    
                    if pd.notna(cantidad) and float(cantidad) > 0:
                        nombre_material = str(col).split('.', 1)[-1].strip().upper()

                        # Capturar postes
                        if "POSTE" in nombre_material:
                            tipo_poste = "CONCRETO" if "CONCRETO" in nombre_material else \
                                        "METALICO" if "METALICO" in nombre_material else \
                                        "FIBRA" if "FIBRA" in nombre_material else "OTRO"

                            # Extraer altura usando regex
                            altura_match = re.search(r'(\d+)\s*MTS?', nombre_material)
                            altura = int(altura_match.group(1)) if altura_match else 0

                            key = f"POSTE|{tipo_poste}|{altura}M"
                            datos[ot]['materiales'][key][nodo] += cantidad
                    
                # SECCIÓN MEJORADA: Procesar códigos N1 y N2 para que sean accesibles para las funciones de mano de obra
                # Procesar código N1
                if pd.notna(codigo_n1) and str(codigo_n1).strip() not in ['', '0', '0.0']:
                    try:
                        potencia_val = 0
                        if pd.notna(potencia_n1) and str(potencia_n1) != "0":
                            try:
                                potencia_val = float(potencia_n1)
                            except:
                                pass
                        
                        if potencia_val == 0:
                            key = "CODIGO 1 LUMINARIA INSTALADA"
                        else:
                            if potencia_val.is_integer():
                                key = f"CODIGO 1 LUMINARIA INSTALADA {int(potencia_val)} W"
                            else:
                                key = f"CODIGO 1 LUMINARIA INSTALADA {potencia_val} W"
                        
                        # Asegurar que el código se almacene como string y en mayúsculas
                        codigo_str = str(codigo_n1).strip().upper()
                        datos[ot]['codigos_n1'][key][nodo].add(codigo_str)
                        
                        # Guardar el código y potencia para este nodo (para usar más tarde)
                        nodos_con_codigos[ot][nodo]['n1'] = {
                            'codigo': codigo_str,
                            'potencia': potencia_val
                        }
                        
                    except Exception as e:
                        # Registrar el error pero continuar con el procesamiento
                        print(f"Error procesando código N1: {str(e)}")
                
                # Procesar código N2
                if pd.notna(codigo_n2) and str(codigo_n2).strip() not in ['', '0', '0.0']:
                    try:
                        potencia_val = 0
                        if pd.notna(potencia_n2) and str(potencia_n2) != "0":
                            try:
                                potencia_val = float(potencia_n2)
                            except:
                                pass
                        
                        if potencia_val == 0:
                            key = "CODIGO 2 LUMINARIA INSTALADA"
                        else:
                            if potencia_val.is_integer():
                                key = f"CODIGO 2 LUMINARIA INSTALADA {int(potencia_val)} W"
                            else:
                                key = f"CODIGO 2 LUMINARIA INSTALADA {potencia_val} W"
                        
                        # Asegurar que el código se almacene como string y en mayúsculas
                        codigo_str = str(codigo_n2).strip().upper()
                        datos[ot]['codigos_n2'][key][nodo].add(codigo_str)
                        
                        # Guardar el código y potencia para este nodo (para usar más tarde)
                        nodos_con_codigos[ot][nodo]['n2'] = {
                            'codigo': codigo_str,
                            'potencia': potencia_val
                        }
                        
                    except Exception as e:
                        # Registrar el error pero continuar con el procesamiento
                        print(f"Error procesando código N2: {str(e)}")
                
                # PROCESAR MATERIALES RETIRADOS
                for (tipo, n), col_codigo in codigo_columns.items():
                    codigo_val = fila.get(col_codigo)
                    col_potencia = None
                    potencia_val = None

                    if tipo in ['LUMINARIA', 'BOMBILLA']:
                        col_potencia = potencia_columns.get((tipo, n), None)
                        if col_potencia:
                            potencia_val = fila.get(col_potencia)
                    
                    codigo_es_valido = (
                        pd.notna(codigo_val)
                        and str(codigo_val).strip() not in ['', '0', '0.0']
                    )
                    
                    potencia_es_valida = False
                    if col_potencia and pd.notna(potencia_val):
                        try: 
                            potencia_float = float(potencia_val)
                            potencia_es_valida = potencia_float > 0
                        except:
                            pass
                    
                    if codigo_es_valido or potencia_es_valida:
                        if tipo == 'FOTOCELDA':
                            entry_name = f"FOTOCELDA RETIRADA {n}"
                        else:
                            potencia_str = ""
                            if potencia_es_valida:
                                potencia_str = (
                                    f"{int(potencia_float)}W"
                                    if potencia_float.is_integer()
                                    else f"{potencia_float}W"
                                )                                               
    
                            if potencia_str:
                                entry_name = f"{tipo} RETIRADA {n} {potencia_str}".strip()
                            else:
                                entry_name = f"{tipo} RETIRADA {n}".strip()
                                
                        key = f"MATERIAL_RETIRADO|{entry_name}"
                        datos[ot]['materiales_retirados'][key][nodo] += 1
                        aspecto = fila["1. Describa Aspectos que Considere se deben tener en cuenta."]
                        if pd.notna(aspecto):
                            aspecto_limpio = str(aspecto).strip().upper()
                            if aspecto_limpio not in ['', 'NA', 'NINGUNO', 'N/A', 'NO', 'NO APLICA']:  # Filtrar valores no válidos
                                datos[ot]['aspectos_retirados'][key][nodo].add(aspecto_limpio)  
                
                # Procesar materiales INSTALADOS
                for mat_col, cant_col in zip(material_cols, cantidad_cols):
                    material = fila[mat_col]
                    cantidad = fila[cant_col]
                    if pd.notna(material) and pd.notna(cantidad) and float(cantidad) > 0.0:
                        material_str = str(material).strip().upper()
                        key = f"MATERIAL|{material_str}"
                        
                        # Convertir cantidad a float y redondear si es necesario
                        cantidad_float = float(cantidad)
                        
                        # Registrar brazos instalados para este nodo
                        if "BRAZO" in material_str:
                            # Extraer longitud del brazo si está disponible
                            longitud_match = re.search(r'(\d+(?:\.\d+)?)\s*M', material_str)
                            longitud = float(longitud_match.group(1)) if longitud_match else 0
                            
                            # Guardar información del brazo para este nodo
                            if nodo not in nodos_con_brazos[ot]:
                                nodos_con_brazos[ot][nodo] = []
                            
                            # Redondear la cantidad si no es un entero
                            cantidad_redondeada = round(cantidad_float)
                            
                            # Guardar información del brazo (descripción, longitud, cantidad)
                            nodos_con_brazos[ot][nodo].append({
                                'descripcion': material_str,
                                'longitud': longitud,
                                'cantidad': cantidad_redondeada if cantidad_redondeada > 0 else 1  # Asegurar al menos 1
                            })
                        
                        # Solo agregar materiales que no sean LUMINARIA N1 o N2
                        if not (material_str.startswith("LUMINARIA N1") or material_str.startswith("LUMINARIA N2")):
                            datos[ot]['materiales'][key][nodo] += cantidad_float
                            datos_por_barrio[barrio_normalizado]['materiales_instalados'][key][ot] += cantidad_float
                            aspecto = fila["1. Describa Aspectos que Considere se deben tener en cuenta."]
                            if pd.notna(aspecto):
                                aspecto_limpio = str(aspecto).strip().upper()
                                if aspecto_limpio not in ['', 'NA', 'NINGUNO', 'N/A', 'NO', 'NO APLICA']:  # Filtrar valores no válidos
                                    datos[ot]['aspectos_materiales'][key][nodo].add(aspecto_limpio)
        
        # Procesar nodos con códigos y brazos para crear entradas de luminarias instaladas
        for ot in datos:
            # Obtener nodos con códigos para esta OT
            nodos_codigos = nodos_con_codigos.get(ot, {})

            # Obtener nodos con brazos para esta OT
            nodos_brazos = nodos_con_brazos.get(ot, {})

            # Diccionario para agrupar luminarias por potencia
            luminarias_por_potencia = defaultdict(int)

            # Procesar cada nodo que tiene código de luminaria
            for nodo, info_codigos in nodos_codigos.items():
                # Verificar si este nodo también tiene brazos instalados
                if nodo in nodos_brazos:
                    # Obtener información de los brazos para este nodo
                    brazos_info = nodos_brazos[nodo]

                    # Calcular la cantidad total de brazos instalados en este nodo
                    total_brazos = sum(brazo['cantidad'] for brazo in brazos_info)

                    # Procesar códigos y potencias
                    for posicion, info in info_codigos.items():
                        codigo = info['codigo']
                        potencia = info['potencia']

                        # Formatear la potencia para la descripción
                        if potencia == 0:
                            potencia_str = "SIN POTENCIA"
                        elif potencia.is_integer():
                            potencia_str = f"{int(potencia)}W"
                        else:
                            potencia_str = f"{potencia}W"

                        # Crear descripción de la luminaria SOLO con la potencia (sin el código)
                        descripcion = f"LUMINARIA CODIGO/BRAZO {potencia_str}"

                        # Crear clave para el material
                        key = f"MATERIAL|{descripcion}"

                        # Determinar la cantidad a instalar (1 por cada brazo)
                        # Si hay múltiples códigos (N1 y N2), distribuir los brazos entre ellos
                        cantidad_a_instalar = 1
                        if len(info_codigos) > 1:
                            # Si hay dos posiciones (N1 y N2), dividir los brazos entre ellas
                            if posicion == 'n1':
                                cantidad_a_instalar = total_brazos // 2 + (total_brazos % 2)  # N1 recibe el extra si es impar
                            else:
                                cantidad_a_instalar = total_brazos // 2
                        else:
                            # Si solo hay un código, asignar todos los brazos a ese código
                            cantidad_a_instalar = total_brazos

                        # Solo agregar si hay brazos para instalar
                        if cantidad_a_instalar > 0:
                            # Agregar la luminaria como material instalado
                            datos[ot]['materiales'][key][nodo] = cantidad_a_instalar

                            # Agregar aspecto si es necesario (aquí sí incluimos el código para referencia)
                            aspecto = f"Luminaria instalada con código: {codigo}"
                            if potencia_str:
                                aspecto += f", potencia: {potencia_str}"
                            aspecto += f", cantidad: {cantidad_a_instalar} brazo(s)"

                            datos[ot]['aspectos_materiales'][key][nodo].add(aspecto)

                            # Actualizar el contador por potencia para estadísticas
                            luminarias_por_potencia[potencia_str] += cantidad_a_instalar

                            # Imprimir información de depuración
                            print(f"INFO: Nodo {nodo} en OT {ot} - Instalada luminaria {posicion.upper()} con potencia {potencia_str}")
                            print(f"      Código: {codigo} (solo para referencia, no se muestra en el material)")
                            print(f"      Cantidad: {cantidad_a_instalar} de {total_brazos} brazos totales")
                            print(f"      Brazos: {', '.join(b['descripcion'] for b in brazos_info)}")

            # Imprimir resumen de luminarias por potencia para esta OT
            if luminarias_por_potencia:
                print(f"\nResumen de luminarias instaladas para OT {ot}:")
                for potencia, cantidad in luminarias_por_potencia.items():
                    print(f"  - LUMINARIA CODIGO/BRAZO {potencia}: {cantidad} unidad(es)")
                        
        return datos, datos_por_barrio, dfs_originales

    except Exception as e:
        logger.error(f"Error procesando {file.filename}: {str(e)}")
        raise HTTPException(500, detail=f"Error en archivo {file.filename}")

def procesar_archivo_mantenimiento(file: UploadFile):
    try:
        contenido = file.file.read()
        xls = pd.ExcelFile(BytesIO(contenido))
        datos = defaultdict(lambda: {
            'nodos': set(),
            'materiales': defaultdict(lambda: defaultdict(int))
        })

        material_pattern = re.compile(r'^MATERIAL\s\d+$', re.IGNORECASE)
        cantidad_pattern = re.compile(r'^CANTIDAD MATERIAL\s\d+$', re.IGNORECASE)

        for hoja in xls.sheet_names:
            df = xls.parse(hoja).rename(columns=lambda x: str(x).strip())
            required_columns = {"6.Nro.Orden Energis", "5.Nodo"}
            if not required_columns.issubset(df.columns):
                continue

            df["5.Nodo"] = df["5.Nodo"].astype(str)
            ot_nodos = df[["6.Nro.Orden Energis", "5.Nodo"]].drop_duplicates()
            for ot, nodo in ot_nodos.itertuples(index=False, name=None):
                datos[ot]['nodos'].add(nodo)

            material_cols = [col for col in df.columns if material_pattern.match(col)]
            cantidad_cols = [col for col in df.columns if cantidad_pattern.match(col)]
            paired_columns = []
            material_nums = {re.search(r'\d+', col).group() for col in material_cols}
            for col in cantidad_cols:
                num = re.search(r'\d+', col).group()
                if num in material_nums:
                    mat_col = next(c for c in material_cols if num in c)
                    paired_columns.append((mat_col, col))

            materiales_data = []
            for mat_col, cant_col in paired_columns:
                temp_df = df[["6.Nro.Orden Energis", "5.Nodo", mat_col, cant_col]].copy()
                temp_df.columns = ["OT", "Nodo", "Material", "Cantidad"]
                materiales_data.append(temp_df)

            combined_df = pd.concat(materiales_data, ignore_index=True)
            combined_df = combined_df[
                (combined_df['Material'].notna()) &
                (~combined_df['Material'].str.strip().str.upper().isin(['', 'NINGUNO', 'NA'])) &
                (pd.to_numeric(combined_df['Cantidad'], errors='coerce') > 0)
            ]
            
            if combined_df.empty:
                continue

            combined_df['Material'] = combined_df['Material'].str.strip().str.upper()
            combined_df['Cantidad'] = pd.to_numeric(combined_df['Cantidad'])
            grouped = combined_df.groupby(['OT', 'Nodo', 'Material'])['Cantidad'].sum().reset_index()

            for row in grouped.itertuples(index=False):
                ot, nodo, material, cantidad = row
                key = f"MATERIAL|{material}"
                datos[ot]['materiales'][key][nodo] += cantidad

        return datos

    except Exception as e:
        logger.error(f"Error procesando {file.filename}: {str(e)}")
        raise HTTPException(500, detail=f"Error en archivo {file.filename}")

def generate_resumen_general(writer, datos_combinados):

    # Recolectar todos los materiales y OTs
    instalados = {}
    retirados = {}
    luminarias = {}
    ots = set()
    
    # Primero identificar todas las OTs
    for ot in datos_combinados.keys():
        ots.add(ot)
    
    # Ordenar OTs para mantener consistencia
    ots_ordenadas = sorted(ots)
    
    # Ahora procesar todos los datos para cada tipo de material
    for ot, info in datos_combinados.items():
        # Proceso para códigos N1
        for key, nodos_data in info.get('codigos_n1', {}).items():
            nombre = f"LUMINARIA {key}"
            if nombre not in luminarias:
                luminarias[nombre] = {
                    'tipo': 'LUMINARIA',
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            
            # Contar el número total de códigos en todos los nodos para esta OT
            codigos_count = sum(len(codigos) for codigos in nodos_data.values())
            luminarias[nombre]['total'] += codigos_count
            luminarias[nombre]['ots'][ot] += codigos_count
        
        # Proceso para códigos N2
        for key, nodos_data in info.get('codigos_n2', {}).items():
            nombre = f"LUMINARIA {key}"
            if nombre not in luminarias:
                luminarias[nombre] = {
                    'tipo': 'LUMINARIA',
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            
            # Contar el número total de códigos en todos los nodos para esta OT
            codigos_count = sum(len(codigos) for codigos in nodos_data.values())
            luminarias[nombre]['total'] += codigos_count
            luminarias[nombre]['ots'][ot] += codigos_count
            
        # Procesar materiales instalados
        for mat_key, nodo_quantities in info['materiales'].items():
            _, nombre = mat_key.split('|', 1)
            total = sum(nodo_quantities.values())
            
            if nombre not in instalados:
                instalados[nombre] = {
                    'tipo': 'INSTALADO',
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            instalados[nombre]['total'] += total
            instalados[nombre]['ots'][ot] += total
        
        # Procesar materiales retirados
        for mat_key, nodo_quantities in info.get('materiales_retirados', {}).items():
            _, nombre = mat_key.split('|', 1)
            total = sum(nodo_quantities.values())
            
            if nombre not in retirados:
                retirados[nombre] = {
                    'tipo': 'RETIRADO',
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            retirados[nombre]['total'] += total
            retirados[nombre]['ots'][ot] += total
        
        transporte_luminarias_total = 0
        transporte_luminarias_por_ot = defaultdict(int)

        # Contar luminarias instaladas y retiradas para transporte
        for nodo in info.get('nodos', []):
            # Contar códigos N1 y N2 válidos
            codigos_count = 0
            for key_codigo, nodos_valores in info.get('codigos_n1', {}).items():
                if nodo in nodos_valores:
                    codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                    codigos_count += len(codigos_validos)

            for key_codigo, nodos_valores in info.get('codigos_n2', {}).items():
                if nodo in nodos_valores:
                    codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                    codigos_count += len(codigos_validos)

            # Contar luminarias en materiales instalados
            luminarias_instaladas = 0
            for mat_key, nodo_quantities in info['materiales'].items():
                if "|" in mat_key:
                    material_name = mat_key.split('|', 1)[1].upper()
                    if ("LUMINARIA" in material_name or "LUM" in material_name or 
                        "LED" in material_name or "LAMP" in material_name or 
                        "FOCO" in material_name) and nodo in nodo_quantities:
                        luminarias_instaladas += nodo_quantities[nodo]

            # Contar luminarias en materiales retirados
            luminarias_retiradas = 0
            for mat_key, nodo_quantities in info.get('materiales_retirados', {}).items():
                if "|" in mat_key:
                    material_name = mat_key.split('|', 1)[1].upper()
                    if ("LUMINARIA" in material_name or "RETIRADA" in material_name or 
                        "LUM" in material_name or "LED" in material_name or 
                        "LAMP" in material_name or "FOCO" in material_name) and nodo in nodo_quantities:
                        luminarias_retiradas += nodo_quantities[nodo]

            # Calcular total para este nodo (limitado a 5 por nodo)
            total_nodo = codigos_count + luminarias_instaladas + luminarias_retiradas
            if total_nodo > 0:
                transporte_luminarias_total += total_nodo
                transporte_luminarias_por_ot[ot] += total_nodo

        # Limitar el total por OT a un máximo razonable

        # Agregar a la sección de instalados si hay transporte
        if transporte_luminarias_total > 0:
            nombre_transporte = "TRANSP.LUMINARIAS, PROYECTORES"
            if nombre_transporte not in instalados:
                instalados[nombre_transporte] = {
                    'tipo': 'INSTALADO',
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            instalados[nombre_transporte]['total'] = transporte_luminarias_total
            for ot_key, cantidad in transporte_luminarias_por_ot.items():
                instalados[nombre_transporte]['ots'][ot_key] = cantidad
        
        
    # Verificación de totales y depuración
    for nombre, info in luminarias.items():
        # Recalcular el total sumando todas las cantidades por OT para confirmar
        total_from_ots = sum(info['ots'].values())
        if info['total'] != total_from_ots:
            # Ajustar si hay discrepancia
            info['total'] = total_from_ots
    
    # Crear DataFrames para cada sección
    columns = ['Nombre Material', 'Unidad', 'Cantidad Total'] + [f'OT_{ot}' for ot in ots_ordenadas]
    
    # Si no hay datos, salir
    if not (luminarias or instalados or retirados):
        return

    # Escribir directamente al Excel sin usar pandas para mayor control
    sheet_name = 'Resumen_general'
    worksheet = writer.book.create_sheet(sheet_name)
    
    # Configuración de estilos
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados de columnas
    for col, header in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    current_row = 2
    
    # Función para escribir una sección
    def write_section(title, data, row):
        # Encabezado de sección
        cell = worksheet.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_alignment
        
        # Combinar celdas para el encabezado
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        
        row += 1
        
        # Escribir datos
        for nombre, info in sorted(data.items()):
            worksheet.cell(row=row, column=1, value=nombre)
            worksheet.cell(row=row, column=2, value=info['unidad'])
            worksheet.cell(row=row, column=3, value=info['total'])
            
            # Cantidades por OT
            for i, ot in enumerate(ots_ordenadas, 4):
                worksheet.cell(row=row, column=i, value=info['ots'].get(ot, 0))
            
            row += 1
            
        return row

    # Escribir sección de luminarias
    if luminarias:
        current_row = write_section("LUMINARIAS INSTALADAS", luminarias, current_row)
        current_row += 1  # Espacio entre secciones
    
    # Escribir sección de materiales instalados
    if instalados:
        current_row = write_section("MATERIALES INSTALADOS", instalados, current_row)
        current_row += 1  # Espacio entre secciones
    
    # Escribir sección de materiales retirados
    if retirados:
        current_row = write_section("MATERIALES RETIRADOS", retirados, current_row)
    
    # Ajustar anchos de columna
    for idx, col in enumerate(columns, 1):
        col_width = len(str(col)) + 2
        # Para la primera columna, usar un ancho mayor
        if idx == 1:
            col_width = 40
        worksheet.column_dimensions[get_column_letter(idx)].width = col_width
    
    # Configurar congelación de paneles
    worksheet.freeze_panes = 'D2'
        
def generate_barrio_sheet(writer, barrio_data, barrio_name):
    
    # Normalizar nombre de hoja
    sheet_name = f"{barrio_name.strip().replace('/', '_')[:25]}_R"[:31]
    
    # Recolectar datos del barrio
    materials = {}
    ots = set()
    
    # Procesar materiales instalados
    for mat_key, ot_quantities in barrio_data['materiales_instalados'].items():
        _, nombre = mat_key.split('|', 1)
        for ot, cantidad in ot_quantities.items():
            ots.add(ot)
            key = (nombre, 'INSTALADO')
            
            if key not in materials:
                materials[key] = {
                    'nombre': nombre,
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            materials[key]['total'] += cantidad
            materials[key]['ots'][ot] += cantidad
    
    # Procesar materiales retirados
    for mat_key, ot_quantities in barrio_data['materiales_retirados'].items():
        _, nombre = mat_key.split('|', 1)
        for ot, cantidad in ot_quantities.items():
            ots.add(ot)
            key = (f"{nombre} (RETIRADO)", 'RETIRADO')
            
            if key not in materials:
                materials[key] = {
                    'nombre': key[0],
                    'unidad': 'UND',
                    'total': 0,
                    'ots': defaultdict(int)
                }
            materials[key]['total'] += cantidad
            materials[key]['ots'][ot] += cantidad

    # Crear DataFrame
    rows = []
    for key, data in materials.items():
        row = {
            'Nombre Material': data['nombre'],
            'Unidad': data['unidad'],
            'Cantidad Total': data['total']
        }
        
        # Agregar columnas por OT
        for ot in ots:
            row[f'OT_{ot}'] = data['ots'].get(ot, 0)
        
        rows.append(row)
    
    if not rows:
        return

    df = pd.DataFrame(rows)
    columns = ['Nombre Material', 'Unidad', 'Cantidad Total'] + sorted([f'OT_{ot}' for ot in ots], key=lambda x: x.split('_')[1])
    df = df[columns]

    # Escribir en Excel
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Formatear hoja
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Ajustar anchos de columna
    for idx, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).map(len).max(),
            len(col))
        worksheet.column_dimensions[get_column_letter(idx)].width = max_len + 2

    # Congelar paneles
    worksheet.freeze_panes = 'D2'

def plantilla_mano_obra(worksheet, df, plantilla):
    """
    Agrega la plantilla de mano de obra a la hoja de Excel.
    
    Args:
        worksheet: Hoja de Excel activa
        df: DataFrame con los datos
        plantilla: Plantilla de mano de obra
    """
    header_font = Font(bold=True)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Determinar posición inicial
    start_row = len(df) + 4  # Usar len(df) en lugar de df.shape[0]
    
    # Escribir encabezados
    headers = ['DESCRIPCION MANO DE OBRA', 'UNIDAD', 'CANTIDAD']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.border = cell_border
    
    # Escribir datos
    for idx, item in enumerate(plantilla, 1):
        row_num = start_row + idx
        descripcion = item['DESCRIPCION MANO DE OBRA']
        
        # Reemplazar "DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA" con "DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA"
        #if descripcion == "DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA":
        #    descripcion = "DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA"
            
        worksheet.cell(row=row_num, column=1, value=descripcion).border = cell_border
        worksheet.cell(row=row_num, column=2, value=item['UNIDAD']).border = cell_border
        worksheet.cell(row=row_num, column=3, value=item['CANTIDAD']).border = cell_border
    
    # Ajustar anchos de columna
    worksheet.column_dimensions['A'].width = 45
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 12

def agregar_tabla_mano_obra(wb, datos, ot, plantilla_mo):
    """
    Agrega una tabla de mano de obra al libro de Excel.
    
    Args:
        wb: Libro de Excel
        datos: Datos procesados
        ot: Número de OT
        plantilla_mo: Plantilla de mano de obra
    
    Returns:
        None
    """
    try:
        # Crear una hoja para la mano de obra
        sheet_name = f"MO_{ot}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Configurar encabezados
        headers = ["DESCRIPCION MANO DE OBRA", "UNIDAD", "CANTIDAD", "NODOS RELACIONADOS"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 80
        
        # Inicializar diccionarios para acumular cantidades y nodos por descripción
        mo_acumulada = {}
        nodos_por_descripcion = {}
        materiales_por_descripcion_nodo = {}
        
        # CAMBIO CRÍTICO: Forzar la aparición de instalación de luminarias SOLO si hay LUMINARIA CODIGO/BRAZO
        # Primero, buscar si hay LUMINARIA CODIGO/BRAZO en los datos
        hay_luminarias_codigo_brazo = False
        nodos_con_luminarias_codigo_brazo = []
    
        # Verificar si hay LUMINARIA CODIGO/BRAZO en materiales instalados
        for material_key, nodos_qty in datos[ot]['materiales'].items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "LUMINARIA CODIGO/BRAZO" in material_name:
                    for nodo, cantidad in nodos_qty.items():
                        if cantidad > 0:
                            hay_luminarias_codigo_brazo = True
                            nodos_con_luminarias_codigo_brazo.append(nodo)
    
        # MEJORA: Determinar si hay brazos grandes o pequeños en la OT para decidir el tipo de instalación
        hay_brazos_grandes = False
        hay_brazos_pequenos = False
        nodos_con_brazos_grandes = []
        nodos_con_brazos_pequenos = []

        # IMPORTANTE: Solo verificar brazos si hay LUMINARIA CODIGO/BRAZO
        if hay_luminarias_codigo_brazo:
            for nodo in datos[ot]['nodos']:
                for material_key, nodos_qty in datos[ot]['materiales'].items():
                    if "|" in material_key:
                        material_name = material_key.split("|")[1].upper()
                        if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                            # Intentar extraer la longitud del brazo
                            longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                            if longitud_match:
                                try:
                                    longitud = int(longitud_match.group(1))
                                    # Considerar brazos de 3 metros o más como grandes (para canasta)
                                    if longitud >= 3:  # Si el brazo es mayor o igual a 3 metros, usar canasta
                                        hay_brazos_grandes = True
                                        nodos_con_brazos_grandes.append(nodo)
                                    else:
                                        hay_brazos_pequenos = True
                                        nodos_con_brazos_pequenos.append(nodo)
                                except:
                                    pass
                            # Si no se puede extraer la longitud pero contiene indicación de tamaño grande
                            elif "2 1/2" in material_name or "2.5" in material_name:
                                hay_brazos_grandes = True
                                nodos_con_brazos_grandes.append(nodo)
                            # Verificar explícitamente si es un brazo de 3 metros
                            elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                                hay_brazos_grandes = True
                                nodos_con_brazos_grandes.append(nodo)
                            else:
                                hay_brazos_pequenos = True
                                nodos_con_brazos_pequenos.append(nodo)
    
        # FORZAR la aparición de instalación de luminarias SOLO si hay LUMINARIA CODIGO/BRAZO
        #if hay_luminarias_codigo_brazo:
        #    # Si hay brazos grandes, forzar instalación en canasta
        #    if hay_brazos_grandes:
        #        descripcion_forzada = "INSTALACION DE LUMINARIAS EN CANASTA"
        #        mo_acumulada[descripcion_forzada] = 0  # Inicializar en 0, se calculará después
        #        nodos_por_descripcion[descripcion_forzada] = []
        #        materiales_por_descripcion_nodo[descripcion_forzada] = {}
        #        
        #        # Contar la cantidad real de LUMINARIA CODIGO/BRAZO en nodos con brazos grandes
        #        for nodo in nodos_con_brazos_grandes:
        #            cantidad_luminarias = 0
        #            for material_key, nodos_qty in datos[ot]['materiales'].items():
        #                if "|" in material_key:
        #                    material_name = material_key.split("|")[1].upper()
        #                    if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
        #                        cantidad_luminarias += nodos_qty[nodo]
        #            
        #            if cantidad_luminarias > 0:
        #                mo_acumulada[descripcion_forzada] += cantidad_luminarias
        #                nodos_por_descripcion[descripcion_forzada].append(nodo)
        #                
        #                # Crear un mensaje explicativo
        #                mensaje = f"LUMINARIA CODIGO/BRAZO CON BRAZOS GRANDES (>= 3M): {cantidad_luminarias} unidades"
        #                
        #                materiales_por_descripcion_nodo[descripcion_forzada][nodo] = {
        #                    'instalados': [mensaje],
        #                    'retirados': []
        #                }

            # Si hay brazos pequeños, forzar instalación en camioneta/escalera
            if hay_brazos_pequenos:
                descripcion_forzada = "INSTALACION DE LUMINARIAS EN CAMIONETA"
                mo_acumulada[descripcion_forzada] = 0  # Inicializar en 0, se calculará después
                nodos_por_descripcion[descripcion_forzada] = []
                materiales_por_descripcion_nodo[descripcion_forzada] = {}
                
                # Contar la cantidad real de LUMINARIA CODIGO/BRAZO en nodos con brazos pequeños
                for nodo in nodos_con_brazos_pequenos:
                    cantidad_luminarias = 0
                    for material_key, nodos_qty in datos[ot]['materiales'].items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                                cantidad_luminarias += nodos_qty[nodo]
                    
                    if cantidad_luminarias > 0:
                        mo_acumulada[descripcion_forzada] += cantidad_luminarias
                        nodos_por_descripcion[descripcion_forzada].append(nodo)
                        
                        # Crear un mensaje explicativo
                        mensaje = f"LUMINARIA CODIGO/BRAZO CON BRAZOS PEQUEÑOS (< 3M): {cantidad_luminarias} unidades"
                        
                        materiales_por_descripcion_nodo[descripcion_forzada][nodo] = {
                            'instalados': [mensaje],
                            'retirados': []
                        }

            # Si no se detectaron brazos específicos pero hay LUMINARIA CODIGO/BRAZO, forzar instalación genérica
            if not hay_brazos_grandes and not hay_brazos_pequenos:
                descripcion_forzada = "INSTALACION DE LUMINARIAS EN CAMIONETA"
                mo_acumulada[descripcion_forzada] = 0  # Inicializar en 0, se calculará después
                nodos_por_descripcion[descripcion_forzada] = []
                materiales_por_descripcion_nodo[descripcion_forzada] = {}
                
                # Contar la cantidad real de LUMINARIA CODIGO/BRAZO en todos los nodos
                for nodo in nodos_con_luminarias_codigo_brazo:
                    cantidad_luminarias = 0
                    for material_key, nodos_qty in datos[ot]['materiales'].items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                                cantidad_luminarias += nodos_qty[nodo]
                    
                    if cantidad_luminarias > 0:
                        mo_acumulada[descripcion_forzada] += cantidad_luminarias
                        nodos_por_descripcion[descripcion_forzada].append(nodo)
                        
                        # Crear un mensaje explicativo
                        mensaje = f"LUMINARIA CODIGO/BRAZO SIN BRAZOS ESPECÍFICOS: {cantidad_luminarias} unidades"
                        
                        materiales_por_descripcion_nodo[descripcion_forzada][nodo] = {
                            'instalados': [mensaje],
                            'retirados': []
                        }
                
                # Si no se encontró ninguna cantidad, asignar al menos 1
                if mo_acumulada[descripcion_forzada] == 0:
                    mo_acumulada[descripcion_forzada] = 1
                    nodos_por_descripcion[descripcion_forzada] = ["FORZADO"]
                    materiales_por_descripcion_nodo[descripcion_forzada] = {
                        "FORZADO": {
                            'instalados': ["LUMINARIA CODIGO/BRAZO DETECTADA"],
                            'retirados': []
                        }
                    }
        
        # Procesar cada nodo normalmente (esto puede agregar más información a la partida forzada)
        for nodo in datos[ot]['nodos']:
            # Obtener el tipo de suelo para este nodo si existe
            tipo_suelo = datos[ot]['tipos_suelo'].get(nodo, "")
            
            # Procesar cada partida de mano de obra
            for partida in plantilla_mo:
                descripcion = partida['DESCRIPCION MANO DE OBRA']
                
                # IMPORTANTE: Para instalación de luminarias, solo procesar si hay LUMINARIA CODIGO/BRAZO
                if "INSTALACION DE LUMINARIAS" in descripcion.upper():
                    # Verificar si hay LUMINARIA CODIGO/BRAZO en este nodo específico
                    tiene_luminaria_codigo_brazo = False
                    for material_key, nodos_qty in datos[ot]['materiales'].items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                tiene_luminaria_codigo_brazo = True
                                break
                    
                    # Si no hay LUMINARIA CODIGO/BRAZO en este nodo, saltar esta partida
                    if not tiene_luminaria_codigo_brazo:
                        continue
                
                # Calcular cantidad de mano de obra para esta partida en este nodo
                # Pasar el tipo de suelo como nuevo parámetro
                cantidad, materiales_relacionados, materiales_retirados = calcular_cantidad_mano_obra(
                        descripcion,
                        datos[ot]['materiales'],
                        datos[ot]['materiales_retirados'],
                        nodo,
                        datos[ot].get('codigos_n1', {}),
                        datos[ot].get('codigos_n2', {}),
                        tipo_suelo,  # AGREGAR COMA AQUÍ
                        # AGREGAR ESTE PARÁMETRO:
                        # Determinar el bloque basado en la descripción
                        "Conexión a tierra" if any(kw in descripcion.upper() for kw in ["CONEXIÓN A CABLE A TIERRA", "INSTALACION KIT SPT", "INSTALACION DE ATERRIZAJES"]) 
                        else "Postes" if any(kw in descripcion.upper() for kw in ["APERTURA", "APLOMADA", "CONCRETADA", "HINCADA", "ALQUILER"]) 
                        else "Otros"
                )
                
                # Si hay cantidad, acumular y registrar nodo
                if cantidad > 0:
                    if descripcion not in mo_acumulada:
                        mo_acumulada[descripcion] = 0
                        nodos_por_descripcion[descripcion] = []
                        materiales_por_descripcion_nodo[descripcion] = {}
                    
                    mo_acumulada[descripcion] += cantidad
                    nodos_por_descripcion[descripcion].append(nodo)
                    
                    # Guardar materiales relacionados para este nodo
                    materiales_por_descripcion_nodo[descripcion][nodo] = {
                        'instalados': materiales_relacionados,
                        'retirados': materiales_retirados
                    }
        
        # ÚLTIMO RECURSO: Si después de todo no hay instalación de luminarias, pero hay LUMINARIA CODIGO/BRAZO, forzarla
        tiene_instalacion_luminarias = False
        for desc in mo_acumulada.keys():
            if "INSTALACION DE LUMINARIAS" in desc.upper():
                tiene_instalacion_luminarias = True
                break
        
        if not tiene_instalacion_luminarias and hay_luminarias_codigo_brazo:
            descripcion_forzada = "INSTALACION DE LUMINARIAS EN CAMIONETA"
            mo_acumulada[descripcion_forzada] = 0  # Inicializar en 0
            nodos_por_descripcion[descripcion_forzada] = []
            materiales_por_descripcion_nodo[descripcion_forzada] = {}
            
            # Contar la cantidad real de LUMINARIA CODIGO/BRAZO en todos los nodos
            for nodo in nodos_con_luminarias_codigo_brazo:
                cantidad_luminarias = 0
                for material_key, nodos_qty in datos[ot]['materiales'].items():
                    if "|" in material_key:
                        material_name = material_key.split("|")[1].upper()
                        if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                            cantidad_luminarias += nodos_qty[nodo]
                
                if cantidad_luminarias > 0:
                    mo_acumulada[descripcion_forzada] += cantidad_luminarias
                    nodos_por_descripcion[descripcion_forzada].append(nodo)
                    
                    # Crear un mensaje explicativo
                    mensaje = f"LUMINARIA CODIGO/BRAZO: {cantidad_luminarias} unidades"
                    
                    materiales_por_descripcion_nodo[descripcion_forzada][nodo] = {
                        'instalados': [mensaje],
                        'retirados': []
                    }
            
            # Si no se encontró ninguna cantidad, asignar al menos 1
            if mo_acumulada[descripcion_forzada] == 0:
                mo_acumulada[descripcion_forzada] = 1
                nodos_por_descripcion[descripcion_forzada] = ["FORZADO"]
                materiales_por_descripcion_nodo[descripcion_forzada] = {
                    "FORZADO": {
                        'instalados': ["LUMINARIA CODIGO/BRAZO DETECTADA"],
                        'retirados': []
                    }
                }
        
        # CORRECCIÓN PARA TRANSP.LUMINARIAS, PROYECTORES: Asegurar que el transporte coincida con la suma de instalaciones y desmontajes
        if "TRANSP.LUMINARIAS, PROYECTORES" in mo_acumulada:
            # Calcular el total de luminarias instaladas y desmontadas
            total_instaladas = 0
            total_desmontadas = 0
            
            # Contar instalaciones
            for desc in mo_acumulada:
                if "INSTALACION DE LUMINARIAS" in desc.upper():
                    total_instaladas += mo_acumulada[desc]
            
            # Contar desmontajes
            for desc in mo_acumulada:
                if "DESMONTAJE DE LUMINARIAS" in desc.upper():
                    total_desmontadas += mo_acumulada[desc]
            
            # El transporte debe ser igual a la suma de instalaciones y desmontajes
            total_esperado = total_instaladas + total_desmontadas
            
            # Actualizar el valor del transporte para que coincida con el total esperado
            mo_acumulada["TRANSP.LUMINARIAS, PROYECTORES"] = total_esperado
                
            # Actualizar la información de materiales para reflejar el ajuste
            for nodo in nodos_por_descripcion["TRANSP.LUMINARIAS, PROYECTORES"]:
                if nodo in materiales_por_descripcion_nodo["TRANSP.LUMINARIAS, PROYECTORES"]:
                    materiales_por_descripcion_nodo["TRANSP.LUMINARIAS, PROYECTORES"][nodo]['instalados'].append(
                        f"TRANSPORTE AJUSTADO PARA COINCIDIR CON INSTALACIONES Y DESMONTAJES (TOTAL: {total_esperado})"
                    )
        
        # Llenar la tabla con los datos acumulados
        row_num = 2
        
        # CAMBIO IMPORTANTE: Primero escribir la instalación de luminarias para asegurar que aparezca
        for partida in plantilla_mo:
            descripcion = partida['DESCRIPCION MANO DE OBRA']
            unidad = partida['UNIDAD']
            
            # Primero procesar solo las partidas de instalación de luminarias
            if "INSTALACION DE LUMINARIAS" in descripcion.upper():
                if descripcion in mo_acumulada and mo_acumulada[descripcion] > 0:
                    cantidad = mo_acumulada[descripcion]
                    nodos = nodos_por_descripcion[descripcion]
                    
                    # Crear texto de nodos relacionados con sus materiales
                    nodos_texto = []
                    codigos_n1 = set()
                    codigos_n2 = set()
                    brazos = set()
                    otros_materiales = set()

                    for nodo in nodos:
                        if nodo in materiales_por_descripcion_nodo[descripcion]:
                            materiales = materiales_por_descripcion_nodo[descripcion][nodo]

                            # Formatear materiales instalados
                            mat_instalados = materiales['instalados']
                            mat_retirados = materiales['retirados']

                            # Extraer códigos N1, N2 y brazos
                            for mat in mat_instalados:
                                if "CÓDIGO N1:" in mat:
                                    codigos = mat.replace("CÓDIGO N1:", "").strip()
                                    codigos_n1.update([c.strip() for c in codigos.split(',')])
                                elif "CÓDIGO N2:" in mat:
                                    codigos = mat.replace("CÓDIGO N2:", "").strip()
                                    codigos_n2.update([c.strip() for c in codigos.split(',')])
                                elif "BRAZO" in mat:
                                    brazos.add(mat)
                                else:
                                    otros_materiales.add(mat)

                    # Crear un resumen consolidado
                    if codigos_n1:
                        nodos_texto.append(f"CÓDIGO N1: {', '.join(sorted(codigos_n1))}")
                    if codigos_n2 and any(c.strip().lower() != "no" for c in codigos_n2):
                        nodos_texto.append(f"CÓDIGO N2: {', '.join(sorted(codigos_n2))}")
                    if brazos:
                        nodos_texto.append(f"BRAZOS: {', '.join(sorted(brazos))}")
                    if otros_materiales:
                        nodos_texto.append(f"OTROS: {', '.join(sorted(otros_materiales))}")
                    
                    # Escribir fila en la tabla
                    ws.cell(row=row_num, column=1).value = descripcion
                    ws.cell(row=row_num, column=2).value = unidad
                    ws.cell(row=row_num, column=3).value = cantidad
                    ws.cell(row=row_num, column=4).value = "\n".join(nodos_texto) if nodos_texto else "Sin detalles"
                    
                    # Aplicar formato
                    for col in range(1, 5):
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    row_num += 1
        
        # Luego procesar el resto de partidas
        for partida in plantilla_mo:
            descripcion = partida['DESCRIPCION MANO DE OBRA']
            unidad = partida['UNIDAD']
            
            # Saltar las partidas de instalación de luminarias que ya se procesaron
            if "INSTALACION DE LUMINARIAS" in descripcion.upper():
                continue
            
            # Si hay cantidad acumulada para esta descripción, mostrarla
            if descripcion in mo_acumulada and mo_acumulada[descripcion] > 0:
                cantidad = mo_acumulada[descripcion]
                nodos = nodos_por_descripcion[descripcion]
                
                # Crear texto de nodos relacionados con sus materiales
                nodos_texto = []
                codigos_n1 = set()
                codigos_n2 = set()
                brazos = set()
                otros_materiales = set()

                for nodo in nodos:
                    if nodo in materiales_por_descripcion_nodo[descripcion]:
                        materiales = materiales_por_descripcion_nodo[descripcion][nodo]

                        # Formatear materiales instalados
                        mat_instalados = materiales['instalados']
                        mat_retirados = materiales['retirados']

                        # Extraer códigos N1, N2 y brazos
                        for mat in mat_instalados:
                            if "CÓDIGO N1:" in mat:
                                codigos = mat.replace("CÓDIGO N1:", "").strip()
                                codigos_n1.update([c.strip() for c in codigos.split(',')])
                            elif "CÓDIGO N2:" in mat:
                                codigos = mat.replace("CÓDIGO N2:", "").strip()
                                codigos_n2.update([c.strip() for c in codigos.split(',')])
                            elif "BRAZO" in mat:
                                brazos.add(mat)
                            else:
                                otros_materiales.add(mat)

                # Crear un resumen consolidado
                if codigos_n1:
                    nodos_texto.append(f"CÓDIGO N1: {', '.join(sorted(codigos_n1))}")
                if codigos_n2 and any(c.strip().lower() != "no" for c in codigos_n2):
                    nodos_texto.append(f"CÓDIGO N2: {', '.join(sorted(codigos_n2))}")
                if brazos:
                    nodos_texto.append(f"BRAZOS: {', '.join(sorted(brazos))}")
                if otros_materiales:
                    nodos_texto.append(f"OTROS: {', '.join(sorted(otros_materiales))}")
                
                # Escribir fila en la tabla
                ws.cell(row=row_num, column=1).value = descripcion
                ws.cell(row=row_num, column=2).value = unidad
                ws.cell(row=row_num, column=3).value = cantidad
                ws.cell(row=row_num, column=4).value = "\n".join(nodos_texto) if nodos_texto else "Sin detalles"
                
                # Aplicar formato
                for col in range(1, 5):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                row_num += 1
        
        # Ajustar altura de filas para acomodar el texto
        for row in ws.iter_rows(min_row=2, max_row=row_num-1):
            max_lines = 1
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    max_lines = max(max_lines, lines)
            
            # Establecer altura basada en el número de líneas (aproximadamente 15 puntos por línea)
            ws.row_dimensions[cell.row].height = max(15, min(max_lines * 15, 409))  # Máximo permitido es 409
        
    except Exception as e:
        logger.error(f"Error agregando tabla de mano de obra: {str(e)}")
        raise

def agregar_hoja_asociaciones(writer, datos_combinados):
    wb = writer.book
    ws = wb.create_sheet("Asociaciones")

    # Estilos
    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red    = PatternFill("solid", fgColor="FF0000")
    gray   = PatternFill("solid", fgColor="CCCCCC")
    bold   = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    # Cargo la plantilla (una sola vez)
    plantilla = cargar_plantilla_mano_obra()

    # Los cuatro bloques con sus keywords
    bloques = [
                    ("Postes",
                    ["APERTURA", "APLOMADA", "CONCRETADA", "HINCADA", "ALQUILER"],  # Removido "BOTADO DE ESCOMBROS"
                    ["POSTE"]),
                    ("Instalación luminarias",
                    ["INSTALACION LUMINARIAS"],
                    ["LUMINARIA", "FOTOCELDA", "GRILLETE", "BRAZO"]),
                    ("Conexión a tierra",
                    ["CONEXIÓN A CABLE A TIERRA", "INSTALACION KIT SPT", "INSTALACION DE ATERRIZAJES", "BOTADO DE ESCOMBROS", "RECUPERACION ZONA DURA"],  # Específico para zona dura
                    ["KIT DE PUESTA A TIERRA", "CONECT PERF", "CONECTOR BIME/COM", "ALAMBRE", "TUERCA", "TORNILLO", "VARILLA"]),
                    ("Desmontaje / Transporte",
                    ["DESMONTAJE", "TRANSPORTE", "TRANSP.", "TRANSPORTE COLLARINES"],
                    ["ALAMBRE", "BRAZO", "CÓDIGO", "CABLE", "ABRAZADERA", "GRILLETE"]),
                    ("Instalación de cables",
                    ["INSTALACION CABLE"],
                    ["CABLE", "TPX", "ALAMBRE"]),
                    ("Otros trabajos",
                    ["VESTIDA CONJUNTO 1 O 2 PERCHAS", "VESTIDA CONJUNTO 3 O MAS PERCHAS", "VESTIDA", "CAJA", "PINTADA", "EXCAVACION", "RECUPERACION ZONA", "SOLDADURA", "INSTALACION TRAMA", "INSTALACION CORAZA"],  # Solo "RECUPERACION ZONA" sin "DURA"
                    ["PERCHA", "CAJA", "TUBO", "TUBERIA", "CONDUIT"])
                ]

    row = 1
    for ot, info in datos_combinados.items():
        # 1) Cabecera de OT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"OT: {ot}")
        c.font = Font(bold=True, size=14)
        c.alignment = center
        row += 2

        # 2) Defino nodos ordenados por fecha de sincronización
        nodos = sorted(
            info.get('fechas_sync', {}).keys(),
            key=lambda n: pd.to_datetime(
                info['fechas_sync'].get(n, ""),
                format='%d/%m/%Y %H:%M:%S',
                errors='coerce'
            )
        )

        # 3) Por cada nodo...
        for nodo in nodos:
            # Cabecera de nodo
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cn = ws.cell(row=row, column=1, value=f"Nodo: {nodo}")
            cn.font = Font(bold=True, size=12)
            cn.fill = gray
            row += 1

            # VERIFICACIÓN: Determinar si hay luminarias o códigos en este nodo
            hay_luminarias = False
            hay_codigos = False

            # Verificar códigos N1 y N2 válidos (excluyendo NO, N/A, NA)
            for _, nodos_valores in info.get('codigos_n1', {}).items():
                if nodo in nodos_valores:
                    # Filtrar códigos NO, N/A, NA
                    codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                    if codigos_validos:
                        hay_codigos = True
                        break
                    
            if not hay_codigos:
                for _, nodos_valores in info.get('codigos_n2', {}).items():
                    if nodo in nodos_valores:
                        # Filtrar códigos NO, N/A, NA
                        codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                        if codigos_validos:
                            hay_codigos = True
                            break
                        
            # Verificar materiales que podrían ser luminarias
            for material_key, nodos_qty in info['materiales'].items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    if ("LUMINARIA" in material_name or "LUM" in material_name or 
                        "LED" in material_name or "LAMP" in material_name or 
                        "FOCO" in material_name):
                        if nodo in nodos_qty and nodos_qty[nodo] > 0:
                            hay_luminarias = True
                            break
            # 4) Los bloques
            for titulo, kw_mo, kw_mat in bloques:
                # Encabezado de bloque
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                cb = ws.cell(row=row, column=1, value=titulo)
                cb.font = Font(bold=True, color="FFFFFF")
                cb.fill = red
                row += 1

                # Cabeceras de mini-tabla
                headers = [
                    "Mano de obra", "Unidad", "Cant. MO",
                    "Material Nuevo", "Cant. Mat.",
                    "Material Ret.",   "Cant. Mat."
                ]
                for col_idx, h in enumerate(headers, start=1):
                    ch = ws.cell(row=row, column=col_idx, value=h)
                    ch.font = bold
                    ch.border = border
                    ch.alignment = center
                    if col_idx == 1:
                        ch.fill = red
                row += 1

                # Filtrar partidas de la plantilla que pertenecen a este bloque
                partidas = [
                    item for item in plantilla
                    if any(k in item['DESCRIPCION MANO DE OBRA'].upper() for k in kw_mo)
                ]

                # CASO ESPECIAL: Si es el bloque de instalación de luminarias y hay luminarias o códigos,
                # forzar la aparición de al menos una partida
                if titulo == "Instalación luminarias" and (hay_luminarias or hay_codigos):
                    # Buscar la partida de instalación en camioneta
                    partida_instalacion = next(
                        (item for item in partidas if "INSTALACION DE LUMINARIAS EN CAMIONETA" in item['DESCRIPCION MANO DE OBRA']),
                        None
                    )
                    
                    if partida_instalacion:
                        desc = partida_instalacion['DESCRIPCION MANO DE OBRA']
                        und = partida_instalacion['UNIDAD']
                        
                        # Crear una lista de materiales relacionados
                        lst_inst = []
                        
                        # Agregar códigos si existen
                        for key_codigo, nodos_valores in info.get('codigos_n1', {}).items():
                            if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                                lst_inst.append(f"CÓDIGO N1: {', '.join(nodos_valores[nodo])}")
                        
                        for key_codigo, nodos_valores in info.get('codigos_n2', {}).items():
                            if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                                lst_inst.append(f"CÓDIGO N2: {', '.join(nodos_valores[nodo])}")
                        
                        # Agregar materiales de luminarias si existen
                        for material_key, nodos_qty in info['materiales'].items():
                            if "|" in material_key:
                                material_name = material_key.split("|")[1].upper()
                                if ("LUMINARIA" in material_name or "LUM" in material_name or 
                                    "LED" in material_name or "LAMP" in material_name or 
                                    "FOCO" in material_name):
                                    if nodo in nodos_qty and nodos_qty[nodo] > 0:
                                        lst_inst.append(f"{material_name} ({nodos_qty[nodo]})")
                        
                        # Si no se encontró nada específico, agregar un mensaje genérico
                        if not lst_inst:
                            lst_inst = ["LUMINARIA INSTALADA (FORZADO)"]
                        
                        # Calcular la suma de materiales instalados
                        sum_inst = 1  # Al menos 1
                        
                        # Escribir la fila forzada
                        ws.cell(row=row, column=1, value=desc).border = border
                        ws.cell(row=row, column=2, value=und).border = border
                        ws.cell(row=row, column=3, value=sum_inst).border = border
                        ws.cell(row=row, column=4, value="; ".join(lst_inst)).border = border
                        ws.cell(row=row, column=5, value=sum_inst).border = border
                        ws.cell(row=row, column=6, value="").border = border
                        ws.cell(row=row, column=7, value="").border = border
                        row += 1

                # Para cada partida, calcular la mano de obra necesaria
                for item in partidas:
                    desc = item['DESCRIPCION MANO DE OBRA']
                    und = item['UNIDAD']
                    
                    # Saltar la partida de instalación en camioneta si ya la forzamos
                    if titulo == "Instalación luminarias" and (hay_luminarias or hay_codigos) and "INSTALACION DE LUMINARIAS EN CAMIONETA" in desc:
                        continue
                    
                    # Obtener el tipo de suelo para este nodo si existe
                    tipo_suelo = info['tipos_suelo'].get(nodo, "")
                    
                    # Usar la función mejorada para calcular cantidades de mano de obra
                    # Pasar el tipo de suelo como parámetro
                    total_mo, lst_inst, lst_ret = calcular_cantidad_mano_obra(
                        desc, 
                        info.get('materiales', {}), 
                        info.get('materiales_retirados', {}),
                        nodo,
                        info.get('codigos_n1', {}),
                        info.get('codigos_n2', {}),
                        tipo_suelo
                    )
                    
                    # Si no hay mano de obra requerida, continuamos con la siguiente partida
                    if total_mo == 0:
                        continue
                    
                    # Contar materiales instalados y retirados
                    # Para materiales instalados, obtener las cantidades reales
                    materiales_inst_con_qty = []
                    for mat_name in lst_inst:
                        for material_key, nodos_qty in info.get('materiales', {}).items():
                            if "|" in material_key and material_key.split("|")[1].upper() == mat_name and nodo in nodos_qty:
                                qty = nodos_qty[nodo]
                                materiales_inst_con_qty.append(f"{mat_name} ({qty})")
                                break
                        else:
                            materiales_inst_con_qty.append(mat_name)
                    
                    # Para materiales retirados, obtener las cantidades reales
                    materiales_ret_con_qty = []
                    for mat_name in lst_ret:
                        for material_key, nodos_qty in info.get('materiales_retirados', {}).items():
                            if "|" in material_key and material_key.split("|")[1].upper() == mat_name and nodo in nodos_qty:
                                qty = nodos_qty[nodo]
                                materiales_ret_con_qty.append(f"{mat_name} ({qty})")
                                break
                        else:
                            materiales_ret_con_qty.append(mat_name)
                    
                    # Calcular sumas totales de materiales
                    sum_inst = sum(
                        nodos_qty[nodo] 
                        for material_key, nodos_qty in info.get('materiales', {}).items() 
                        if "|" in material_key and material_key.split("|")[1].upper() in lst_inst and nodo in nodos_qty
                    ) if lst_inst else 0
                    
                    sum_ret = sum(
                        nodos_qty[nodo] 
                        for material_key, nodos_qty in info.get('materiales_retirados', {}).items() 
                        if "|" in material_key and material_key.split("|")[1].upper() in lst_ret and nodo in nodos_qty
                    ) if lst_ret else 0

                    # Escribo la fila
                    ws.cell(row=row, column=1, value=desc).border = border
                    ws.cell(row=row, column=2, value=und).border = border
                    ws.cell(row=row, column=3, value=total_mo).border = border

                    ws.cell(row=row, column=4, value="; ".join(materiales_inst_con_qty)).border = border
                    ws.cell(row=row, column=5, value=sum_inst if sum_inst > 0 else "").border = border
                    
                    ws.cell(row=row, column=6, value="; ".join(materiales_ret_con_qty)).border = border
                    ws.cell(row=row, column=7, value=sum_ret if sum_ret > 0 else "").border = border
                    row += 1

                row += 1  # espacio tras bloque

            row += 2  # espacio tras nodo

        row += 4  # espacio tras OT

    # Ajusto ancho de columnas
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

def cargar_plantilla_mano_obra():
    """
    Define manualmente la plantilla de mano de obra en lugar de cargarla desde un archivo Excel.
    Incluye todas las partidas específicas para luminarias según el tamaño de los brazos.
    
    Returns:
        list: Lista de diccionarios con las partidas de mano de obra
    """
    try:
        # Definir manualmente todas las partidas de mano de obra
        plantilla_manual = [
            # Partidas de instalación de luminarias
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE LUMINARIAS EN CAMIONETA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE LUMINARIAS EN CANASTA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE LUMINARIAS HORIZONTAL ADOSADA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            
            # Partidas de desmontaje de luminarias
            {
                'DESCRIPCION MANO DE OBRA': "DESMONTAJE DE LUMINARIAS EN CAMIONETA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "DESMONTAJE DE LUMINARIAS EN CANASTA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            
            # Partidas de transporte
            {
                'DESCRIPCION MANO DE OBRA': "TRANSP.LUMINARIAS, PROYECTORES",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE DE BRAZOS 1 1/2\" HASTA 3 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE DE BRAZOS 2 1/2\" HASTA 6 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE DE CABLE",
                'UNIDAD': "ML",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE VARILLA TIERRA Y/O KIT TIERRA ACERO",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE DE TAPA PARA CAJA DE A.P",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE PERCHA CON AISLADOR",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSPORTE COLLARINES",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            
            # Partidas de postes
            {
                'DESCRIPCION MANO DE OBRA': "TRANSP.POSTE.CONC.12MT.SITIO SIN INCREME",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSP.POSTE.CONC.18MT.SITIO SIN INCREME",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "TRANSP.POSTE.METALICO DE 4 A 12MT",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APERTURA HUECOS POSTES ANCLAS SECUNDARIAS DE 8 A 10 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APERTURA HUECOS POSTES ANCLAS PRIMARIA DE 11 A 14",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APLOMADA POSTES DE CONCRETO DE 8 A 10 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APLOMADA POSTES DE CONCRETO DE 11 A 14 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APLOMADA POSTES METALICOS Y/O FIBRA VIDRIO 8 A 10",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "APLOMADA POSTES METALICOS Y/O FIBRA VIDRIO 11 A 14",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "CONCRETADA DE POSTE CONCRETO DE 8 A 12 M INCLUYE MATERIALES Y MO",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "CONCRETADA DE POSTE PRIMARIOS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "CONCRETADA DE POSTE METALICO 8 A 12 MT INCLUYE MATERIALES Y MO",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "CONCRETADA DE POSTE FIBRA 8 A 12 MT INCLUYE MATERIALES",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "CONCRETADA ENTRE POSTE Y CAJA INCLUYE MATERIALES E INSTALACIÓN",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTES CONCRETO DE 8 A 12 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTES DE 14 MTS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTE METALICO DE 4 A 8M",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTE METALICO DE 10 A 12M",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTE FIBRA DE 8M",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "HINCADA DE POSTE FIBRA DE 10 A 12M",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            
            # Partidas de conexión a tierra
            {
                'DESCRIPCION MANO DE OBRA': "CONEXIÓN A CABLE A TIERRA (INSTALACION CONECTOR DE SPT)",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION KIT SPT CON CINTA METALICA",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE ATERRIZAJES SECUNDARIOS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
        'DESCRIPCION MANO DE OBRA': "RECUPERACION ZONA DURA",
        'UNIDAD': "UND",
        'CANTIDAD': 0
    },
    {
        'DESCRIPCION MANO DE OBRA': "BOTADO DE ESCOMBROS",
        'UNIDAD': "UND",
        'CANTIDAD': 0
    },
            
            # Partidas de cables
            {
                'DESCRIPCION MANO DE OBRA': "DESMONTAJE CABLE SECUNDARIO #4 A #2/0",
                'UNIDAD': "ML",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION CABLE SECUNDARIO #4 A #2/0 AEREO",
                'UNIDAD': "ML",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION CABLE SECUNDARIO #4 A #2/0 SUBTERRANEO",
                'UNIDAD': "ML",
                'CANTIDAD': 0
            },
            
            # Partidas de cajas
            {
                'DESCRIPCION MANO DE OBRA': "CAJA DE A.P 0,4X0,4 MT INCLUYE MATERIALES E INSTALACION",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            
            # Partidas adicionales
            {
                'DESCRIPCION MANO DE OBRA': "ALQUILER DE EQUIPO \"MACHINE COMPRESOR\"",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION TRAMA DE SEGURIDAD",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "VESTIDA CONJUNTO 1 O 2 PERCHAS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "VESTIDA CONJUNTO 3 O MAS PERCHAS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "PINTADA DE NODOS",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "EXCAVACION",
                'UNIDAD': "M3",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "RECUPERACION ZONA",
                'UNIDAD': "M2",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "SOLDADURA POR PUNTO",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION CORAZA",
                'UNIDAD': "ML",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE CUBIERTA GEL",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            },
            {
                'DESCRIPCION MANO DE OBRA': "INSTALACION DE EMPALME",
                'UNIDAD': "UND",
                'CANTIDAD': 0
            }
        ]
        
        return plantilla_manual
        
    except Exception as e:
        logger.error(f"Error creando plantilla manual: {str(e)}")
        return []  

def calcular_cantidad_mano_obra(descripcion, materiales_instalados, materiales_retirados, nodo, codigos_n1=None, codigos_n2=None, tipo_suelo=None, bloque_actual=None):
    """
    Calcula la cantidad de mano de obra necesaria para una partida específica en un nodo.
    
    Args:
        descripcion: Descripción de la partida de mano de obra
        materiales_instalados: Diccionario de materiales instalados
        materiales_retirados: Diccionario de materiales retirados
        nodo: Nodo actual
        codigos_n1: Diccionario de códigos N1 (opcional)
        codigos_n2: Diccionario de códigos N2 (opcional)
    
    Returns:
        tuple: (cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados)
    """    
    
    cantidad_mo = 0
    materiales_instalados_relacionados = []
    materiales_retirados_relacionados = []
    
    # Convertir descripción a mayúsculas para comparaciones insensibles a mayúsculas/minúsculas
    descripcion_upper = descripcion.upper()
    
    # Verificamos si hay diccionarios válidos para evitar errores
    if codigos_n1 is None:
        codigos_n1 = {}
    if codigos_n2 is None:
        codigos_n2 = {}
    mano_obra_por_nodo = {}
    
    # ======== CASO ESPECIAL: RECUPERACION ZONA DURA ========
    if "RECUPERACION ZONA DURA" in descripcion_upper:
        # Verificar si el tipo de suelo es "Zona Dura"
        if tipo_suelo and "ZONA DURA" in tipo_suelo.upper():
            # Verificar si hay trabajos que requieran recuperación de zona
            hay_trabajo_con_postes = False
            hay_excavacion = False
            hay_kit_puesta_tierra = False

            # Buscar KIT DE PUESTA A TIERRA en materiales instalados
            for material_key, nodos_qty in materiales_instalados.items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_kit_puesta_tierra = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # Buscar postes instalados en este nodo
            for key, nodos_qty in materiales_instalados.items():
                if "|" in key:
                    material_name = key.split("|")[1].upper()
                    if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_trabajo_con_postes = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # Si no hay postes instalados, buscar postes retirados
            if not hay_trabajo_con_postes:
                for key, nodos_qty in materiales_retirados.items():
                    if "|" in key:
                        material_name = key.split("|")[1].upper()
                        if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                            hay_trabajo_con_postes = True
                            materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            break
                        
            # Buscar si hay excavación en este nodo
            for key, nodos_qty in materiales_instalados.items():
                if "|" in key:
                    material_name = key.split("|")[1].upper()
                    if ("EXCAVACION" in material_name or "ZANJA" in material_name) and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_excavacion = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # CORRECCIÓN: Si hay kit de puesta a tierra O trabajo con postes O excavación, asignar recuperación
            if hay_kit_puesta_tierra or hay_trabajo_con_postes or hay_excavacion:
                cantidad_mo = 1 * 0.3  # Asignar 1 unidad por nodo
                materiales_instalados_relacionados.append(f"TIPO DE SUELO: ZONA DURA")
                return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: RECUPERACION ZONA ========
    if "RECUPERACION ZONA" in descripcion_upper and "DURA" not in descripcion_upper:
        # Verificar si el tipo de suelo NO es "Zona Dura" (es decir, es zona blanda)
        es_zona_blanda = False
        if tipo_suelo:
            if "ZONA BLANDA" in tipo_suelo.upper():
                es_zona_blanda = True
            elif "ZONA DURA" not in tipo_suelo.upper():
                # Si no especifica "Zona Dura", asumir que es zona blanda
                es_zona_blanda = True
        else:
            # Si no hay información de tipo de suelo, asumir que es zona blanda
            es_zona_blanda = True

        if es_zona_blanda:
            # Verificar si hay trabajos que requieran recuperación de zona
            hay_trabajo_con_postes = False
            hay_excavacion = False
            hay_kit_puesta_tierra = False

            # Buscar KIT DE PUESTA A TIERRA en materiales instalados
            for material_key, nodos_qty in materiales_instalados.items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_kit_puesta_tierra = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # Buscar postes instalados en este nodo
            for key, nodos_qty in materiales_instalados.items():
                if "|" in key:
                    material_name = key.split("|")[1].upper()
                    if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_trabajo_con_postes = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # Si no hay postes instalados, buscar postes retirados
            if not hay_trabajo_con_postes:
                for key, nodos_qty in materiales_retirados.items():
                    if "|" in key:
                        material_name = key.split("|")[1].upper()
                        if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                            hay_trabajo_con_postes = True
                            materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            break
                        
            # Buscar si hay excavación en este nodo
            for key, nodos_qty in materiales_instalados.items():
                if "|" in key:
                    material_name = key.split("|")[1].upper()
                    if ("EXCAVACION" in material_name or "ZANJA" in material_name) and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_excavacion = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
            # CORRECCIÓN: Si hay kit de puesta a tierra O trabajo con postes O excavación, asignar recuperación
            if hay_kit_puesta_tierra or hay_trabajo_con_postes or hay_excavacion:
                cantidad_mo = 1 * 0.3  # Asignar 1 unidad por nodo
                materiales_instalados_relacionados.append(f"TIPO DE SUELO: ZONA BLANDA")
                return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados       
   
    # ======== CASO ESPECIAL: CONEXIÓN A CABLE A TIERRA (INSTALACION CONECTOR DE SPT) ========
    if "CONEXIÓN A CABLE A TIERRA" in descripcion_upper or "INSTALACION CONECTOR DE SPT" in descripcion_upper:
        # Verificar si este nodo NO tiene un KIT SPT instalado
        tiene_kit_spt = False
        
        # Buscar KIT DE PUESTA A TIERRA en materiales instalados para este nodo
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_kit_spt = True
                    break
        
        # Si NO tiene KIT SPT, asignar esta partida
        if not tiene_kit_spt:
            # Verificar si hay algún material eléctrico que requiera conexión a tierra
            hay_material_electrico = False
            
            # Buscar materiales eléctricos que típicamente requieren conexión a tierra
            for material_key, nodos_qty in materiales_instalados.items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    # Verificar si hay luminarias, cables, u otros componentes eléctricos
                    if any(keyword in material_name for keyword in ["LUMINARIA", "CABLE", "LAMPARA", "LED", "FOCO", "CONECTOR"]) and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        hay_material_electrico = True
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
            
            # Si hay material eléctrico o códigos N1/N2, asignar la partida
            hay_codigos = False
            for _, nodos_valores in codigos_n1.items():
                if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                    hay_codigos = True
                    break
            
            if not hay_codigos:
                for _, nodos_valores in codigos_n2.items():
                    if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                        hay_codigos = True
                        break
            
            if hay_material_electrico or hay_codigos:
                cantidad_mo = 1
                if not materiales_instalados_relacionados:
                    materiales_instalados_relacionados.append("CONEXIÓN A TIERRA SIN KIT SPT ESPECÍFICO")
                return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: TRANSPORTE COLLARINES ========
    if "TRANSPORTE COLLARINES" in descripcion_upper:
        cantidad_total = 0
        tiene_abrazadera = False
        tiene_grillete = False
        
        # Buscar ABRAZADERA CIEGA 1 CARA GALV 6" en materiales instalados
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es una abrazadera ciega
                if "ABRAZADERA CIEGA 1 CARA GALV 6" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_abrazadera = True
                    cantidad_total += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                
                # Verificar si es un grillete galvanizado
                if "GRILLETE GALVANIZADO 1 1/2" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_grillete = True
                    cantidad_total += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # También buscar en materiales retirados
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es una abrazadera ciega
                if "ABRAZADERA CIEGA 1 CARA GALV 6" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_abrazadera = True
                    cantidad_total += nodos_qty[nodo]
                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                
                # Verificar si es un grillete galvanizado
                if "GRILLETE GALVANIZADO 1 1/2" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_grillete = True
                    cantidad_total += nodos_qty[nodo]
                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Si se encontró al menos uno de los materiales, asignar mano de obra
        if tiene_abrazadera or tiene_grillete:
            cantidad_mo = cantidad_total
            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: INSTALACION KIT SPT CON CINTA METALICA ========
    if "INSTALACION KIT SPT CON CINTA METALICA" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar específicamente KIT DE PUESTA A TIERRA en materiales instalados
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es el kit de puesta a tierra
                if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    cantidad_total += qty
                    materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: DESMONTAJE CABLE SECUNDARIO #4 A #2/0 ========
    if "DESMONTAJE CABLE SECUNDARIO #4 A #2/0" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar específicamente CABLE AL #4 o CABLE TRENZADO 2x4 en materiales retirados
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es uno de los cables específicos mencionados
                es_cable_al4 = "CABLE AL #4" in material_name
                es_cable_trenzado_2x4 = "CABLE TRENZADO 2X4" in material_name
                es_cable_tpx = "CABLE TPX 2X4 AWG XLPE + 48.69 AAAC" in material_name
                es_cable_al_tpx = "CABLE AL TPX 2X2+1X2 AWG" in material_name
                
                if (es_cable_al4 or es_cable_trenzado_2x4 or es_cable_tpx or es_cable_al_tpx) and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    cantidad_total += qty
                    materiales_retirados_relacionados.append(f"{material_name} ({qty})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: INSTALACION CABLE SECUNDARIO #4 A #2/0 AEREO ========
    if "INSTALACION CABLE SECUNDARIO #4 A #2/0 AEREO" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar específicamente CABLE TPX 2x4 AWG XLPE + 48.69 AAAC y CABLE AL TPX 2x2+1x2 AWG
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es uno de los cables específicos mencionados
                es_cable_tpx = "CABLE TPX 2X4 AWG XLPE + 48.69 AAAC" in material_name
                es_cable_al_tpx = "CABLE AL TPX 2X2+1X2 AWG" in material_name
                
                if (es_cable_tpx or es_cable_al_tpx) and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    # Multiplicar por 3 la cantidad para la mano de obra
                    cantidad_total += qty * 3
                    materiales_instalados_relacionados.append(f"{material_name} ({qty} x 3 = {qty*3})")
                # También considerar otros cables que puedan requerir esta mano de obra
                elif any(kw in material_name for kw in ["CABLE", "ALAMBRE", "TPX", "CONDUCTOR"]) and "#4" in material_name and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    cantidad_total += qty
                    materiales_instalados_relacionados.append(f"{material_name} ({qty})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: VESTIDA CONJUNTO 1 O 2 PERCHAS ========
    if "VESTIDA CONJUNTO" in descripcion_upper and "PERCHAS" in descripcion_upper and "1 O 2" in descripcion_upper:
        # Verificar si hay PERCHA GALV 1 PUESTO o PERCHA GALV 2 PUESTOS en este nodo
        tiene_percha = False
        cantidad_perchas = 0
        
        # Lista para almacenar los tipos de perchas encontrados
        perchas_encontradas = []
    
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Buscar perchas de 1 o 2 puestos con diferentes variantes de escritura
                es_percha_1_puesto = "PERCHA GALV 1 PUESTO" in material_name
                es_percha_2_puestos = ("PERCHA GALV 2 PUESTO" in material_name or 
                                      "PERCHA GALV 2 PUESTOS" in material_name)
                
                if (es_percha_1_puesto or es_percha_2_puestos) and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_percha = True
                    cantidad_perchas += nodos_qty[nodo]
                    perchas_encontradas.append(material_name)
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
    
        # También buscar en materiales retirados por si hay perchas que se están reemplazando
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Buscar perchas de 1 o 2 puestos con diferentes variantes de escritura
                es_percha_1_puesto = "PERCHA GALV 1 PUESTO" in material_name
                es_percha_2_puestos = ("PERCHA GALV 2 PUESTO" in material_name or 
                                      "PERCHA GALV 2 PUESTOS" in material_name)
                
                if (es_percha_1_puesto or es_percha_2_puestos) and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    # No sumamos a cantidad_perchas porque no queremos duplicar la mano de obra
                    # Solo registramos para información
                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
    
        if tiene_percha:
            cantidad_mo = cantidad_perchas
            # Agregar resumen de lo encontrado
            if perchas_encontradas:
                materiales_instalados_relacionados.append(f"Total perchas 1-2 puestos: {cantidad_perchas}")
            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: VESTIDA CONJUNTO 3 O MAS PERCHAS ========
    if "VESTIDA CONJUNTO" in descripcion_upper and "PERCHAS" in descripcion_upper and "3 O MAS" in descripcion_upper:
        # Verificar si hay perchas de 3 o más puestos en este nodo
        tiene_percha_grande = False
        cantidad_perchas = 0
        
        # Lista para almacenar los tipos de perchas encontrados
        perchas_encontradas = []
    
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Buscar perchas de 3 o más puestos
                if "PERCHA GALV" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    # Intentar extraer el número de puestos
                    puestos_match = re.search(r'(\d+)\s*PUESTO', material_name)
                    if puestos_match:
                        try:
                            puestos = int(puestos_match.group(1))
                            if puestos >= 3:
                                tiene_percha_grande = True
                                cantidad_perchas += nodos_qty[nodo]
                                perchas_encontradas.append(material_name)
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        except:
                            pass
                    # Si no se puede extraer el número pero contiene "3" o más
                    elif any(f"{i} PUESTO" in material_name or f"{i} PUESTOS" in material_name for i in range(3, 10)):
                        tiene_percha_grande = True
                        cantidad_perchas += nodos_qty[nodo]
                        perchas_encontradas.append(material_name)
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
    
        # También buscar en materiales retirados por si hay perchas que se están reemplazando
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Buscar perchas de 3 o más puestos
                if "PERCHA GALV" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    # Intentar extraer el número de puestos
                    puestos_match = re.search(r'(\d+)\s*PUESTO', material_name)
                    if puestos_match:
                        try:
                            puestos = int(puestos_match.group(1))
                            if puestos >= 3:
                                # No sumamos a cantidad_perchas porque no queremos duplicar la mano de obra
                                # Solo registramos para información
                                materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        except:
                            pass
                    # Si no se puede extraer el número pero contiene "3" o más
                    elif any(f"{i} PUESTO" in material_name or f"{i} PUESTOS" in material_name for i in range(3, 10)):
                        materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
    
        if tiene_percha_grande:
            cantidad_mo = cantidad_perchas
            # Agregar resumen de lo encontrado
            if perchas_encontradas:
                materiales_instalados_relacionados.append(f"Total perchas 3+ puestos: {cantidad_perchas}")
            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CASO ESPECIAL: TRANSPORTE PERCHA CON AISLADOR ========
    if "TRANSPORTE PERCHA" in descripcion_upper and "AISLADOR" in descripcion_upper:
        # Verificar si hay AISLADOR CARR A.P ANSI 53-2- 3" y cualquier tipo de PERCHA GALV en este nodo
        tiene_aislador = False
        tiene_percha_galv = False
        cantidad_aisladores = 0
        cantidad_perchas = 0

        # Lista para almacenar los tipos de perchas encontrados para mostrar en materiales relacionados
        perchas_encontradas = []
        aisladores_encontrados = []

        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()

                # Detectar aisladores
                if "AISLADOR CARR A.P ANSI 53-2" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_aislador = True
                    cantidad_aisladores += nodos_qty[nodo]
                    aisladores_encontrados.append(f"{material_name} ({nodos_qty[nodo]})")

                # Detectar todos los tipos de perchas galvanizadas
                if "PERCHA GALV" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_percha_galv = True
                    cantidad_perchas += nodos_qty[nodo]
                    perchas_encontradas.append(f"{material_name} ({nodos_qty[nodo]})")

                    # Verificar específicamente los diferentes tipos de perchas
                    es_percha_1_puesto = "PERCHA GALV 1 PUESTO" in material_name
                    es_percha_2_puestos = "PERCHA GALV 2 PUESTO" in material_name or "PERCHA GALV 2 PUESTOS" in material_name
                    es_percha_3_puestos = "PERCHA GALV 3 PUESTO" in material_name or "PERCHA GALV 3 PUESTOS" in material_name

                    # Registrar el tipo específico de percha encontrada
                    if es_percha_1_puesto or es_percha_2_puestos or es_percha_3_puestos:
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # También buscar en materiales retirados por si hay perchas o aisladores que se están transportando
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()

                # Detectar aisladores retirados
                if "AISLADOR CARR A.P ANSI 53-2" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_aislador = True
                    cantidad_aisladores += nodos_qty[nodo]
                    aisladores_encontrados.append(f"{material_name} ({nodos_qty[nodo]})")
                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

                # Detectar todos los tipos de perchas galvanizadas retiradas
                if "PERCHA GALV" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_percha_galv = True
                    cantidad_perchas += nodos_qty[nodo]
                    perchas_encontradas.append(f"{material_name} ({nodos_qty[nodo]})")

                    # Verificar específicamente los diferentes tipos de perchas
                    es_percha_1_puesto = "PERCHA GALV 1 PUESTO" in material_name
                    es_percha_2_puestos = "PERCHA GALV 2 PUESTO" in material_name or "PERCHA GALV 2 PUESTOS" in material_name
                    es_percha_3_puestos = "PERCHA GALV 3 PUESTO" in material_name or "PERCHA GALV 3 PUESTOS" in material_name

                    # Registrar el tipo específico de percha encontrada
                    if es_percha_1_puesto or es_percha_2_puestos or es_percha_3_puestos:
                        materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # Si hay aisladores y perchas, asignar mano de obra para transporte
        if tiene_aislador and tiene_percha_galv:
            # La cantidad debe ser el mínimo entre aisladores y perchas
            cantidad_mo = min(cantidad_aisladores, cantidad_perchas)

            # Si no se encontraron materiales específicos pero sabemos que hay perchas y aisladores
            if not materiales_instalados_relacionados and not materiales_retirados_relacionados:
                materiales_instalados_relacionados.append("TRANSPORTE DE PERCHAS Y AISLADORES")

            # Agregar resumen de lo encontrado
            if aisladores_encontrados:
                materiales_instalados_relacionados.append(f"Total aisladores: {cantidad_aisladores}")
            if perchas_encontradas:
                materiales_instalados_relacionados.append(f"Total perchas: {cantidad_perchas}")

            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    if "INSTALACION DE ATERRIZAJES SECUNDARIOS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar VARILLA COOPERWELD en materiales instalados
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                # Verificar si es una varilla cooperweld
                if ("VARILLA COOPERWELD" in material_name and "5/8" in material_name and 
                    nodo in nodos_qty and nodos_qty[nodo] > 0):
                    cantidad_total += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Si se encontró al menos una varilla, asignar mano de obra
        if cantidad_total > 0:
            cantidad_mo = cantidad_total
            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
                
    # ======== CASO ESPECIAL: TRANSPORTE VARILLA TIERRA Y/O KIT TIERRA ACERO ========
    if "TRASPORTE VARILLA TIERRA" in descripcion_upper or "TRANSPORTE VARILLA TIERRA" in descripcion_upper:
        # Verificar si hay VARILLA COOPERWELD 5/8″ X 2.4 MT o KIT DE PUESTA A TIERRA en este nodo
        tiene_varilla = False
        tiene_kit_tierra = False
        cantidad_varillas = 0
        cantidad_kits = 0
        
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                
                if "VARILLA COOPERWELD" in material_name and "5/8" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_varilla = True
                    cantidad_varillas += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                
                if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_kit_tierra = True
                    cantidad_kits += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Si hay varilla o kit de tierra, asignar la mano de obra
        if tiene_varilla or tiene_kit_tierra:
            # La cantidad debe ser la suma de varillas y kits
            cantidad_mo = cantidad_varillas + cantidad_kits
            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CÁLCULO COMÚN DE LUMINARIAS INSTALADAS ========
    # Calcular la cantidad total de luminarias instaladas (para usar en múltiples secciones)
    luminarias_instaladas_count = 0
    
    luminaria_keywords = [
        "LUMINARIA", "LAMPARA", "LED", "BOMBILLA", "PROYECTOR", 
        "REFLECTOR", "FOCO", "BALASTRO", "CODIGO 1", "CODIGO 2",
        "CODIGO N1", "CODIGO N2", "CODIGO DE LUMINARIA", "LUM"
    ]
    
    # 1. Verificar códigos N1 y N2 instalados
    codigos_encontrados = False
    if codigos_n1:  # Verificar que no sea None
        for key_codigo, nodos_valores in codigos_n1.items():
            if nodo in nodos_valores:
                # Filtrar códigos NO, N/A, NA
                codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                cantidad_codigos = len(codigos_validos)
                if cantidad_codigos > 0:
                    luminarias_instaladas_count += cantidad_codigos
                    codigos_encontrados = True
                    # Agregar códigos a materiales relacionados
                    materiales_instalados_relacionados.append(f"CÓDIGO N1: {', '.join(codigos_validos)}")
    
    if codigos_n2:  # Verificar que no sea None
        for key_codigo, nodos_valores in codigos_n2.items():
            if nodo in nodos_valores:
                # Filtrar códigos NO, N/A, NA
                codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                cantidad_codigos = len(codigos_validos)
                if cantidad_codigos > 0:
                    luminarias_instaladas_count += cantidad_codigos
                    codigos_encontrados = True
                    # Agregar códigos a materiales relacionados
                    materiales_instalados_relacionados.append(f"CÓDIGO N2: {', '.join(codigos_validos)}")
    
    # 2. Buscar luminarias instaladas en materiales
    luminarias_encontradas = False
    for material_key, nodos_qty in materiales_instalados.items():
        if "|" in material_key:
            material_name = material_key.split("|")[1].upper()
            
            # Detección ampliada de luminarias con patrones más flexibles
            es_luminaria = (
                "LUMINARIA" in material_name or 
                "LUM" in material_name or
                "LED" in material_name or
                "LAMP" in material_name or
                "FOCO" in material_name or
                re.search(r'LUMINARIA N[12]', material_name) is not None or
                any(kw in material_name for kw in luminaria_keywords)
            )
            
            if es_luminaria and nodo in nodos_qty:
                qty = nodos_qty[nodo]
                if qty > 0:
                    luminarias_instaladas_count += qty
                    luminarias_encontradas = True
                    materiales_instalados_relacionados.append(f"{material_name} ({qty})")
    
    # ======== INSTALACIÓN DE LUMINARIAS ========
    
    # Verificar si la descripción corresponde a instalación de luminarias
    if "INSTALACION DE LUMINARIAS" in descripcion_upper:
        # CAMBIO CRÍTICO: Buscar EXCLUSIVAMENTE luminarias CODIGO/BRAZO en materiales instalados
        luminarias_codigo_brazo = 0
        tiene_brazos_grandes_reales = False

        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    luminarias_codigo_brazo += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # Si no hay LUMINARIA CODIGO/BRAZO, retornar 0 inmediatamente
        if luminarias_codigo_brazo == 0:
            return 0, [], []

        # VERIFICAR SIEMPRE si hay brazos >= 3M en este nodo (independientemente de si hay LUMINARIA CODIGO/BRAZO)
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                    if longitud_match:
                        try:
                            longitud = int(longitud_match.group(1))
                            if longitud >= 3:
                                tiene_brazos_grandes_reales = True
                                materiales_instalados_relacionados.append(f"BRAZO >= 3M: {material_name} ({nodos_qty[nodo]})")
                                break
                        except:
                            pass
                    # Verificar patrones específicos de brazos grandes
                    elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                        tiene_brazos_grandes_reales = True
                        materiales_instalados_relacionados.append(f"BRAZO 3M: {material_name} ({nodos_qty[nodo]})")
                        break

        # LÓGICA ESTRICTA: Solo asignar el tipo correcto según los brazos
        if "CANASTA" in descripcion_upper:
            if not tiene_brazos_grandes_reales:
                return 0, [], []  # No hay brazos >= 3M, no usar canasta
            else:
                cantidad_mo = luminarias_codigo_brazo  # Usar canasta solo si hay brazos >= 3M
        elif "CAMIONETA" in descripcion_upper:
            if tiene_brazos_grandes_reales:
                return 0, [], []  # Hay brazos >= 3M, no usar camioneta
            else:
                cantidad_mo = luminarias_codigo_brazo  # Usar camioneta solo si NO hay brazos >= 3M

        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    
    # Verificar si la descripción corresponde a desmontaje de luminarias    
    if "DESMONTAJE DE LUMINARIAS" in descripcion_upper and ("CAMIONETA" in descripcion_upper or "CANASTA" in descripcion_upper):
        # Lista ampliada de palabras clave para detectar luminarias retiradas
        #luminaria_keywords = [
        #    "LUMINARIA", "LAMPARA", "LED", "BOMBILLA", "PROYECTOR", 
        #    "REFLECTOR", "FOCO", "BALASTRO", "RETIRADA"
        #]
        luminaria_keywords = [
            "LUMINARIA"
        ]
        
        # Inicializar contador de luminarias retiradas
        luminarias_retiradas_count = 0
        
        # Buscar luminarias retiradas
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                # Verificar si es una luminaria retirada usando múltiples criterios
                es_luminaria_retirada = (
                    any(kw in material_name for kw in luminaria_keywords) or 
                    #"RETIRADA" in material_name or
                    "LUMINARIA RETIRADA" in material_name 
                    #"BOMBILLA RETIRADA" in material_name or
                    #"FOTOCELDA RETIRADA" in material_name
                )
                
                if es_luminaria_retirada and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    luminarias_retiradas_count += qty
                    materiales_retirados_relacionados.append(f"{material_name} ({qty})")
        
        # NUEVO: Determinar si se debe usar canasta o camioneta para el desmontaje
        # basado en el tamaño de los brazos retirados
        usar_canasta_para_desmontaje = False
        brazos_grandes_retirados = []
        brazos_pequenos_retirados = []
        
        # Buscar brazos retirados y determinar su tamaño
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    # Intentar extraer la longitud del brazo
                    longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                    if longitud_match:
                        try:
                            longitud = int(longitud_match.group(1))
                            # CORRECCIÓN: Considerar brazos de 3 metros o más como grandes (para canasta)
                            if longitud >= 3:  # DEBE SER >= 3, no > 3
                                usar_canasta_para_desmontaje = True
                                brazos_grandes_retirados.append(f"{material_name} ({nodos_qty[nodo]})")
                                materiales_retirados_relacionados.append(f"BRAZO >= 3M: {material_name} ({nodos_qty[nodo]})")
                            else:
                                brazos_pequenos_retirados.append(f"{material_name} ({nodos_qty[nodo]})")
                                materiales_retirados_relacionados.append(f"BRAZO < 3M: {material_name} ({nodos_qty[nodo]})")
                        except:
                            pass
                    # Si no se puede extraer la longitud pero contiene indicación de tamaño grande
                    elif "2 1/2" in material_name or "2.5" in material_name:
                        usar_canasta_para_desmontaje = True
                        brazos_grandes_retirados.append(f"{material_name} ({nodos_qty[nodo]})")
                        materiales_retirados_relacionados.append(f"BRAZO GRANDE: {material_name} ({nodos_qty[nodo]})")
                    # Verificar explícitamente si es un brazo de 3 metros o más
                    elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                        usar_canasta_para_desmontaje = True
                        brazos_grandes_retirados.append(f"{material_name} ({nodos_qty[nodo]})")
                        materiales_retirados_relacionados.append(f"BRAZO 3M: {material_name} ({nodos_qty[nodo]})")
                    else:
                        brazos_pequenos_retirados.append(f"{material_name} ({nodos_qty[nodo]})")
                        materiales_retirados_relacionados.append(f"BRAZO: {material_name} ({nodos_qty[nodo]})")
        
        # Asignar la mano de obra según el tipo de desmontaje
        # IMPORTANTE: Solo asignar a uno de los tipos para evitar duplicación
        if "DESMONTAJE DE LUMINARIAS EN CANASTA" in descripcion_upper:
            if usar_canasta_para_desmontaje:
                cantidad_mo = luminarias_retiradas_count
                # Agregar información sobre los brazos grandes encontrados
                if brazos_grandes_retirados:
                    materiales_retirados_relacionados.append(f"DESMONTAJE CON CANASTA POR BRAZOS GRANDES: {', '.join(brazos_grandes_retirados)}")
        elif "DESMONTAJE DE LUMINARIAS EN CAMIONETA" in descripcion_upper:
            if not usar_canasta_para_desmontaje:
                cantidad_mo = luminarias_retiradas_count
                # Agregar información sobre los brazos pequeños encontrados
                if brazos_pequenos_retirados:
                    materiales_retirados_relacionados.append(f"DESMONTAJE CON CAMIONETA POR BRAZOS PEQUEÑOS: {', '.join(brazos_pequenos_retirados)}")
                # Si no hay brazos pero hay luminarias, indicar que se usa camioneta por defecto
                elif luminarias_retiradas_count > 0:
                    materiales_retirados_relacionados.append("DESMONTAJE CON CAMIONETA (NO SE DETECTARON BRAZOS)")
        
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    
    # ======== TRANSP.LUMINARIAS, PROYECTORES ========    
    if "TRANSP.LUMINARIAS, PROYECTORES" in descripcion_upper:
        # NUEVA LÓGICA MEJORADA: Calcular transporte de luminarias priorizando CODIGO/BRAZO

        # 1. Buscar luminarias CODIGO/BRAZO en materiales instalados
        luminarias_codigo_brazo = 0
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    luminarias_codigo_brazo += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # 2. Contar códigos N1 y N2 válidos (excluyendo NO, N/A, NA) si no hay CODIGO/BRAZO
        codigos_cantidad = 0
        if luminarias_codigo_brazo == 0:
            # Contar códigos N1 válidos
            if codigos_n1:
                for key_codigo, nodos_valores in codigos_n1.items():
                    if nodo in nodos_valores:
                        # Filtrar códigos NO, N/A, NA
                        codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                        codigos_cantidad += len(codigos_validos)
                        if len(codigos_validos) > 0:
                            materiales_instalados_relacionados.append(f"CÓDIGO N1: {', '.join(codigos_validos)}")

            # Contar códigos N2 válidos
            if codigos_n2:
                for key_codigo, nodos_valores in codigos_n2.items():
                    if nodo in nodos_valores:
                        # Filtrar códigos NO, N/A, NA
                        codigos_validos = [c for c in nodos_valores[nodo] if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                        codigos_cantidad += len(codigos_validos)
                        if len(codigos_validos) > 0:
                            materiales_instalados_relacionados.append(f"CÓDIGO N2: {', '.join(codigos_validos)}")

        # 3. Contar luminarias instaladas en materiales (si no hay CODIGO/BRAZO ni códigos)
        luminarias_instaladas_count = 0
        if luminarias_codigo_brazo == 0 and codigos_cantidad == 0:
            #luminaria_keywords = [
            #    "LUMINARIA", "LAMPARA", "LED", "BOMBILLA", "PROYECTOR", 
            #    "REFLECTOR", "FOCO", "BALASTRO", "LUM"
            #]
            luminaria_keywords = [ "LUMINARIA", "LUM"
            ]
            for material_key, nodos_qty in materiales_instalados.items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    es_luminaria = any(kw in material_name for kw in luminaria_keywords)
                    if es_luminaria and "CODIGO/BRAZO" not in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        luminarias_instaladas_count += nodos_qty[nodo]
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # 4. Contar luminarias retiradas
        luminarias_retiradas_count = 0
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                es_luminaria_retirada = (
                #    any(kw in material_name for kw in ["LUMINARIA", "LAMPARA", "LED", "BOMBILLA", "PROYECTOR", 
                #                                      "REFLECTOR", "FOCO", "BALASTRO", "LUM"]) or 
                #    "RETIRADA" in material_name
                #)
                     any(kw in material_name for kw in ["LUMINARIA", "LUM"]) )
                
                if es_luminaria_retirada and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    luminarias_retiradas_count += nodos_qty[nodo]
                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # 5. Verificar si hay brazos en el nodo (solo si no hay CODIGO/BRAZO ni códigos ni luminarias)
        tiene_brazos = False
        brazos_encontrados = []
        cantidad_brazos = 0

        if luminarias_codigo_brazo == 0 and codigos_cantidad == 0 and luminarias_instaladas_count == 0:
            for material_key, nodos_qty in materiales_instalados.items():
                if "|" in material_key:
                    material_name = material_key.split("|")[1].upper()
                    if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        tiene_brazos = True
                        cantidad_brazos += nodos_qty[nodo]
                        brazos_encontrados.append(f"{material_name} ({nodos_qty[nodo]})")
                        materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        # LÓGICA MEJORADA: Determinar cantidad de transporte con prioridad para CODIGO/BRAZO
        cantidad_instaladas = 0

        # Prioridad 1: Si hay luminarias CODIGO/BRAZO, usar esa cantidad
        if luminarias_codigo_brazo > 0:
            cantidad_instaladas = luminarias_codigo_brazo
            materiales_instalados_relacionados.append(f"TRANSPORTE POR LUMINARIAS CODIGO/BRAZO: {luminarias_codigo_brazo}")
        # Prioridad 2: Si hay códigos, usar esa cantidad
        #elif codigos_cantidad > 0:
        #    cantidad_instaladas = codigos_cantidad
        #    materiales_instalados_relacionados.append(f"TRANSPORTE POR CÓDIGOS: {codigos_cantidad}")
        ## Prioridad 3: Si hay luminarias en materiales instalados, usar esa cantidad
        #elif luminarias_instaladas_count > 0:
        #    cantidad_instaladas = luminarias_instaladas_count
        #    materiales_instalados_relacionados.append(f"TRANSPORTE POR LUMINARIAS INSTALADAS: {luminarias_instaladas_count}")
        ## Prioridad 4: Si hay brazos pero no hay códigos ni luminarias explícitas, inferir por brazos
        #elif tiene_brazos:
        #    cantidad_instaladas = cantidad_brazos
        #    materiales_instalados_relacionados.append(f"TRANSPORTE POR BRAZOS: {cantidad_brazos}")
#
        # Para retiradas, usar directamente la cantidad calculada
        cantidad_retiradas = luminarias_retiradas_count
        if cantidad_retiradas > 0:
            materiales_retirados_relacionados.append(f"TRANSPORTE DE LUMINARIAS RETIRADAS: {luminarias_retiradas_count}")

        # Asignar la cantidad final
        cantidad_mo = cantidad_instaladas + cantidad_retiradas

        # Agregar información de resumen
        if cantidad_instaladas > 0:
            materiales_instalados_relacionados.append(f"Total luminarias instaladas para transporte: {cantidad_instaladas}")

        if cantidad_retiradas > 0:
            materiales_retirados_relacionados.append(f"Total luminarias retiradas para transporte: {cantidad_retiradas}")

        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados        
    # ======== TRANSPORTE DE BRAZOS ========
    
    # Transporte de brazos de 1 1/2" hasta 3 mts
    if "TRANSPORTE DE BRAZOS 1 1/2\" HASTA 3 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar brazos instalados que coincidan con la descripción
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty:
                    # Verificar si es un brazo de 1 1/2" hasta 3 mts
                    if "1 1/2" in material_name or "1.5" in material_name:
                        # Verificar longitud si está especificada
                        longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                        if longitud_match:
                            try:
                                longitud = int(longitud_match.group(1))
                                if longitud <= 3:
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            except:
                                # Si no se puede extraer la longitud pero parece ser un brazo pequeño
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        else:
                            # Si no hay longitud específica, asumir que está en el rango
                            cantidad_total += nodos_qty[nodo]
                            materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Buscar brazos retirados que coincidan con la descripción
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty:
                    # Verificar si es un brazo de 1 1/2" hasta 3 mts
                    if "1 1/2" in material_name or "1.5" in material_name:
                        # Verificar longitud si está especificada
                        longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                        if longitud_match:
                            try:
                                longitud = int(longitud_match.group(1))
                                if longitud <= 3:
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            except:
                                # Si no se puede extraer la longitud pero parece ser un brazo pequeño
                                cantidad_total += nodos_qty[nodo]
                                materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        else:
                            # Si no hay longitud específica, asumir que está en el rango
                            cantidad_total += nodos_qty[nodo]
                            materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Transporte de brazos 2 1/2" hasta 6 mts
    if "TRANSPORTE DE BRAZOS 2 1/2\" HASTA 6 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar brazos instalados que coincidan con la descripción
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty:
                    # Verificar si es un brazo de 2 1/2" hasta 6 mts
                    if "2 1/2" in material_name or "2.5" in material_name:
                        # Verificar longitud si está especificada
                        longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                        if longitud_match:
                            try:
                                longitud = int(longitud_match.group(1))
                                if 3 < longitud <= 6:
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            except:
                                # Si no se puede extraer la longitud pero parece ser un brazo grande
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        else:
                            # Si no hay longitud específica, asumir que está en el rango
                            cantidad_total += nodos_qty[nodo]
                            materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Buscar brazos retirados que coincidan con la descripción
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "BRAZO" in material_name and nodo in nodos_qty:
                    # Verificar si es un brazo de 2 1/2" hasta 6 mts
                    if "2 1/2" in material_name or "2.5" in material_name:
                        # Verificar longitud si está especificada
                        longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                        if longitud_match:
                            try:
                                longitud = int(longitud_match.group(1))
                                if 3 < longitud <= 6:
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            except:
                                # Si no se puede extraer la longitud pero parece ser un brazo grande
                                cantidad_total += nodos_qty[nodo]
                                materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        else:
                            # Si no hay longitud específica, asumir que está en el rango
                            cantidad_total += nodos_qty[nodo]
                            materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== TRANSPORTE DE CABLES ========
    
    if "TRANSPORTE DE CABLE" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar cables instalados
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if any(kw in material_name for kw in ["CABLE", "ALAMBRE", "TPX", "CONDUCTOR"]) and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    cantidad_total += qty
                    materiales_instalados_relacionados.append(f"{material_name} ({qty})")
        
        # Buscar cables retirados
        for material_key, nodos_qty in materiales_retirados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if any(kw in material_name for kw in ["CABLE", "ALAMBRE", "TPX", "CONDUCTOR"]) and nodo in nodos_qty:
                    qty = nodos_qty[nodo]
                    cantidad_total += qty
                    materiales_retirados_relacionados.append(f"{material_name} ({qty})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== TRANSPORTE DE POSTES ========
    
    # Transporte de postes metálicos (incluye postes de fibra) de 4 a 12 metros
    if "TRANSP.POSTE.METALICO DE 4 A 12MT" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes metálicos o de fibra en el rango de 4 a 12 metros
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                # Detectar postes metálicos o de fibra con criterios más amplios
                if "POSTE" in material_name and nodo in nodos_qty:
                    es_metalico_o_fibra = (
                        "METALICO" in material_name or 
                        "FIBRA" in material_name or
                        "METAL" in material_name
                    )

                    if es_metalico_o_fibra:
                        # Extraer altura del nombre del poste
                        altura_match = re.search(r'(\d+)\s*M', material_name)
                        if altura_match:
                            try:
                                altura = int(altura_match.group(1))
                                if 4 <= altura <= 12:
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                            except:
                                # Si no se puede extraer la altura pero contiene indicación de tamaño
                                if any(f"{i}M" in material_name for i in range(4, 13)):
                                    cantidad_total += nodos_qty[nodo]
                                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        else:
                            # Si no hay altura específica, asumir que está en el rango
                            cantidad_total += nodos_qty[nodo]
                            materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")

        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Transporte de postes de concreto 12 metros
    if "TRANSP.POSTE.CONC.12MT.SITIO SIN INCREME" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto de 12 metros instalados
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    # Verificar si es de 12 metros
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if altura == 12:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        except:
                            # Si no se puede extraer la altura pero contiene "12M"
                            if "12M" in material_name or "12 M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                    else:
                        # Si no hay altura específica pero es de concreto, verificar si podría ser de 12m
                        if "12" in material_name:
                            cantidad_total += nodos_qty[nodo]
                            materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        # Si no se encontraron postes específicos pero hay postes de concreto, forzar al menos 1
        if cantidad_total == 0:
            hay_postes_concreto = False
            for key, nodos_qty in materiales_instalados.items():
                if "|" in key and "POSTE" in key.split("|")[1].upper() and "CONCRETO" in key.split("|")[1].upper() and sum(nodos_qty.values()) > 0:
                    hay_postes_concreto = True
                    break
                
            if hay_postes_concreto:
                cantidad_total = 1
                materiales_instalados_relacionados.append("TRANSPORTE DE POSTE CONCRETO (FORZADO)")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Transporte de postes de concreto 18 metros
    if "TRANSP.POSTE.CONC.18MT.SITIO SIN INCREME" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto de 18 metros instalados
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    # Verificar si es de 18 metros
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if altura == 18 or (altura >= 15 and altura <= 20):  # Rango más amplio
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        except:
                            # Si no se puede extraer la altura pero contiene "18M"
                            if "18M" in material_name or "18 M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                    else:
                        # Si no hay altura específica pero es de concreto, verificar si podría ser de 18m
                        if "18" in material_name:
                            cantidad_total += nodos_qty[nodo]
                            materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
        
        cantidad_mo = cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== POSTES Y OTROS ELEMENTOS ========
    
    # ======== CÓDIGO PARA POSTES DE CONCRETO ========    
    
    # Transporte de postes de concreto 18 metros
    if "TRANSP.POSTE.CONC.18MT.SITIO SIN INCREME" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto de 18 metros instalados
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and "18" in material_name and nodo in nodos_qty:
                    cantidad_total += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Aplomada de postes de concreto (8 a 10 metros)
    if "APLOMADA POSTES DE CONCRETO DE 8 A 10 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 10:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "8M" o "10M"
                            if "8M" in material_name or "9M" in material_name or "10M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Aplomada de postes de concreto (11 a 14 metros)
    if "APLOMADA POSTES DE CONCRETO DE 11 A 14 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 11 <= altura <= 14:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "12M" o "14M"
                            if "11M" in material_name or "12M" in material_name or "13M" in material_name or "14M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Apertura de huecos para postes secundarias (8 a 10 metros)
    if "APERTURA HUECOS POSTES ANCLAS SECUNDARIAS DE 8 A 10 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 10:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "8M" o "10M"
                            if "8M" in material_name or "9M" in material_name or "10M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Apertura de huecos para postes primaria (11 a 14 metros)
    if "APERTURA HUECOS POSTES ANCLAS PRIMARIA DE 11 A 14 MTS" in descripcion_upper or "APERTURA HUECOS POSTES 11 MT A 14 MT Y ANCLAS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 11 <= altura <= 14:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "12M" o "14M"
                            if "11M" in material_name or "12M" in material_name or "13M" in material_name or "14M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Base de concreto para poste 9 mts
    if "BASE DE CONCRETO PARA POSTE 9 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes cercanos a 9 metros
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 10:  # Considerar postes cercanos a 9 metros
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "9M"
                            if "8M" in material_name or "9M" in material_name or "10M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Concretada de poste de concreto de 8 a 12 metros
    if "CONCRETADA DE POSTE CONCRETO DE 8 A 12 M INCLUYE MATERIALES Y MO" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto en ese rango
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 12:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            pass
        
        cantidad_mo += cantidad_total
    
    # Alquiler machine neumático y compresor
    if "ALQUILER DE EQUIPO \"MACHINE COMPRESOR\"" in descripcion_upper or "ALQUILER MACHINE NEUMATICO Y COMPRESOR" in descripcion_upper:
        # Determinar si se necesita compresor basado en trabajos en el nodo
        requiere_compresor = False

        # Buscar instalaciones que típicamente requieren compresor
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                # Postes de concreto típicamente requieren compresor
                if "POSTE" in material_name and "CONCRETO" in material_name and nodo in nodos_qty:
                    requiere_compresor = True
                    materiales_instalados_relacionados.append(material_name)
                    break
                
        # Si se identifica la necesidad de compresor, contabilizar una vez por nodo
        if requiere_compresor:
            cantidad_mo = 1

        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # Hincada de postes de concreto de 14 metros
    if "HINCADA DE POSTES CONCRETO DE 14 MTS" in descripcion_upper:
        cantidad_total = 0
        
        # Buscar postes de concreto de 14 metros
        for key in materiales_instalados:
            if "POSTE" in key.upper() and "CONCRETO" in key.upper() and "14" in key and nodo in materiales_instalados[key]:
                cantidad_total += materiales_instalados[key][nodo]
                materiales_instalados_relacionados.append(key.replace("MATERIAL|", ""))
            # También considerar postes de concreto donde la altura sea exactamente 14
            elif "POSTE" in key.upper() and "CONCRETO" in key.upper() and nodo in materiales_instalados[key]:
                altura_match = re.search(r'(\d+)\s*M', key.upper())
                if altura_match:
                    try:
                        altura = int(altura_match.group(1))
                        if altura == 14:
                            cantidad_total += materiales_instalados[key][nodo]
                            materiales_instalados_relacionados.append(key.replace("MATERIAL|", ""))
                    except:
                        pass
        
        cantidad_mo += cantidad_total
    
    # Botado de escombros
    if "BOTADO DE ESCOMBROS" in descripcion_upper:
        # Variables para controlar qué tipo de trabajo requiere botado de escombros
        tiene_kit_puesta_tierra = False
        tiene_trabajo_con_postes = False

        # 1. Verificar si hay KIT DE PUESTA A TIERRA
        for material_key, nodos_qty in materiales_instalados.items():
            if "|" in material_key:
                material_name = material_key.split("|")[1].upper()
                if "KIT DE PUESTA A TIERRA" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_kit_puesta_tierra = True
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                    break  # Solo necesitamos saber que existe

        # 2. Verificar si hay trabajo con postes (instalados o retirados)
        # Buscar postes instalados
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                    tiene_trabajo_con_postes = True
                    materiales_instalados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                    break

        # Si no hay postes instalados, buscar postes retirados
        if not tiene_trabajo_con_postes:
            for key, nodos_qty in materiales_retirados.items():
                if "|" in key:
                    material_name = key.split("|")[1].upper()
                    if "POSTE" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                        tiene_trabajo_con_postes = True
                        materiales_retirados_relacionados.append(f"{material_name} ({nodos_qty[nodo]})")
                        break
                    
        # 3. CORRECCIÓN: Asignar botado de escombros si hay kit de puesta a tierra O trabajo con postes
        if tiene_kit_puesta_tierra or tiene_trabajo_con_postes:
            cantidad_mo = 1 * 0.1
            if tiene_kit_puesta_tierra:
                materiales_instalados_relacionados.append("BOTADO DE ESCOMBROS POR: KIT PUESTA A TIERRA")
            if tiene_trabajo_con_postes:
                materiales_instalados_relacionados.append("BOTADO DE ESCOMBROS POR: TRABAJO CON POSTES")

            return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    # ======== CÓDIGO PARA POSTES DE FIBRA ========

    # Transporte de postes metálicos (incluye postes de fibra) de 4 a 12 metros    

    # Aplomada de postes metálicos o fibra (8 a 10 metros)
    if "APLOMADA POSTES METALICOS Y/O FIBRA VIDRIO 8 A 10" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes de fibra en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                # Detectar postes de fibra o metálicos
                if ("POSTE FIBRA" in material_name or "POSTE METALICO" in material_name) and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 10:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "8M" o "10M"
                            if "8M" in material_name or "9M" in material_name or "10M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)

        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Aplomada de postes metálicos o fibra (11 a 14 metros)
    if "APLOMADA POSTES METALICOS Y/O FIBRA VIDRIO 11 A 14" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes de fibra en ese rango de altura
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                # Detectar postes de fibra o metálicos
                if ("POSTE FIBRA" in material_name or "POSTE METALICO" in material_name) and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 11 <= altura <= 14:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene alturas específicas
                            if "11M" in material_name or "12M" in material_name or "13M" in material_name or "14M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)

        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Concretada de poste de fibra de 8 a 12 metros
    if "CONCRETADA DE POSTE FIBRA 8 A 12 MT INCLUYE MATERIAL" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes de fibra en ese rango
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE FIBRA" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 8 <= altura <= 12:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene dimensiones apropiadas
                            if "8M" in material_name or "9M" in material_name or "10M" in material_name or "11M" in material_name or "12M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)

        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Hincada de poste de fibra de 8M
    if "HINCADA DE POSTE FIBRA DE 8M" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes de fibra de 8 metros
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE FIBRA" in material_name and "8M" in material_name and nodo in nodos_qty:
                    cantidad_total += nodos_qty[nodo]
                    materiales_instalados_relacionados.append(material_name)
                elif "POSTE FIBRA" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if altura == 8:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            pass
                        
        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Hincada de poste de fibra de 10 a 12M
    if "HINCADA DE POSTE FIBRA DE 10 A 12M" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes de fibra de 10 a 12 metros
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "POSTE FIBRA" in material_name and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 10 <= altura <= 12:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "10M" o "12M"
                            if "10M" in material_name or "11M" in material_name or "12M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)

        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Concretada entre poste y caja
    if "CONCRETADA ENTRE POSTE Y CAJA INCLUYE MATERIALES E INSTALACIÓN" in descripcion_upper:
        cantidad_total = 0

        # Verificar si hay postes de fibra o metálicos instalados
        postes_encontrados = False
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if ("POSTE FIBRA" in material_name or "POSTE METALICO" in material_name) and nodo in nodos_qty:
                    postes_encontrados = True
                    materiales_instalados_relacionados.append(material_name)
                    break
                
        # Verificar si hay cajas instaladas
        cajas_encontradas = False
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if "CAJA" in material_name and nodo in nodos_qty:
                    cajas_encontradas = True
                    materiales_instalados_relacionados.append(material_name)
                    break
                
        # Si hay ambos elementos, asignar mano de obra
        if postes_encontrados and cajas_encontradas:
            cantidad_mo += 1

        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

    # Base para poste metálico tipo Keison
    if "BASE PARA POSTE METALICO TIPO KEISON DE 10 A 12 MT" in descripcion_upper:
        cantidad_total = 0

        # Buscar postes metálicos o de fibra en el rango adecuado
        for key, nodos_qty in materiales_instalados.items():
            if "|" in key:
                material_name = key.split("|")[1].upper()
                if ("POSTE METALICO" in material_name or "POSTE FIBRA" in material_name) and nodo in nodos_qty:
                    # Extraer altura del nombre del poste
                    altura_match = re.search(r'(\d+)\s*M', material_name)
                    if altura_match:
                        try:
                            altura = int(altura_match.group(1))
                            if 10 <= altura <= 12:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)
                        except:
                            # Si no se puede extraer la altura pero contiene "10M" o "12M"
                            if "10M" in material_name or "11M" in material_name or "12M" in material_name:
                                cantidad_total += nodos_qty[nodo]
                                materiales_instalados_relacionados.append(material_name)

        cantidad_mo += cantidad_total
        return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados
    
    # ======== CÓDIGO PARA TUBERÍA CONDUFLEX ========
    
    # Caja de AP e instalación relacionada con tubería conduflex
    if "CAJA DE A.P 0,4X0,4 MT INCLUYE MATERIALES E INSTALACION" in descripcion_upper:
        cantidad_total = 0
        
        # Verificar si hay tubería conduflex en el nodo
        for key in materiales_instalados:
            if "TUBERIA CONDUFLEX" in key.upper() and nodo in materiales_instalados[key]:
                # Contar una caja por nodo donde haya tubería
                cantidad_total = 1
                materiales_instalados_relacionados.append(key.replace("MATERIAL|", ""))
                break
        
        cantidad_mo += cantidad_total
    
    # Devolver la cantidad de mano de obra y los materiales relacionados
    return cantidad_mo, materiales_instalados_relacionados, materiales_retirados_relacionados

def generar_excel(datos_combinados, datos_por_barrio_combinados, dfs_originales_combinados):
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        temp_sheet_name = "PlantillaInicial"
        writer.book.create_sheet(temp_sheet_name)
        sheets_created = False        
        
        generate_resumen_general(writer, datos_combinados)
        
        generate_resumen_tecnicos(writer, datos_combinados, dfs_originales_combinados)  
        
        # IMPORTANTE: Habilitar la generación de la hoja de asociaciones
        #agregar_hoja_asociaciones(writer, datos_combinados) 
                                 
        if datos_combinados:                        
            
            # IMPORTANTE: Generar hojas de mano de obra ANTES de las hojas de OT
            # Esto asegura que las hojas de mano de obra se creen incluso si hay un error después
            #plantilla_mo = cargar_plantilla_mano_obra()
            #for ot, info in datos_combinados.items():
            #    # FORZAR la generación de la hoja de mano de obra para cada OT
            #    try:
            #        # Verificar si hay códigos N1 o N2 en esta OT
            #        hay_codigos = False
            #        for _, nodos_valores in info.get('codigos_n1', {}).items():
            #            if any(len(codigos) > 0 for codigos in nodos_valores.values()):
            #                hay_codigos = True
            #                break
            #        
            #        if not hay_codigos:
            #            for _, nodos_valores in info.get('codigos_n2', {}).items():
            #                if any(len(codigos) > 0 for codigos in nodos_valores.values()):
            #                    hay_codigos = True
            #                    break
            #        
            #        # Si hay códigos, imprimir información de depuración
            #        if hay_codigos:
            #            print(f"INFO: OT {ot} tiene códigos N1/N2. Generando hoja de mano de obra.")
            #        
            #        # Generar la hoja de mano de obra independientemente de si hay códigos o no
            #        agregar_tabla_mano_obra(writer.book, datos_combinados, ot, plantilla_mo)
            #        print(f"INFO: Hoja de mano de obra generada para OT {ot}")
            #    except Exception as e:
            #        print(f"ERROR al generar hoja de mano de obra para OT {ot}: {str(e)}")
            
            for ot, info in datos_combinados.items():
                seen = set()
                nodos_con_fechas = [(nodo, pd.to_datetime(info['fechas_sync'].get(nodo, "01/01/1900 00:00:00"), 
                   format='%d/%m/%Y %H:%M:%S', errors='coerce')) 
                   for nodo in info['nodos'] if nodo not in seen and not seen.add(nodo)]
                nodos_con_fechas.sort(key=lambda x: x[1])  # Sort by timestamp
                nodos_ordenados = [nodo for nodo, _ in nodos_con_fechas]
                num_nodos = len(nodos_ordenados)
                # Obtener postes desde la fila "Nodos postes"
                postes = [nodo.split('_')[0] for nodo in nodos_ordenados]  # Ej: ['1417541', '1417542']
                # Columnas con formato Nodo_X (Poste_Y)
                columnas = [
                    'OT', 
                    'Unidad', 
                    'Cantidad Total',
                    'Fecha Sincronización',  # Esta columna ahora solo tendrá contenido para observaciones
                    *[f"Nodo_{i+1}" for i in range(num_nodos)]
                ]
                filas = []
                
                # Fila OT y Nodos (con postes reales)
                filas.append([ot, '', '', ''] + [''] * num_nodos)
                filas.append(['Nodos postes', '', '', ''] + postes)  # Mostrar postes sin sufijo
                
                # Agregar fila con fechas de sincronización de cada nodo
                fechas_fila = ['Fechas Sincronización', '', '', '']
                for nodo in nodos_ordenados:
                    fecha = info['fechas_sync'].get(nodo, "Sin fecha")
                    fechas_fila.append(fecha)
                filas.append(fechas_fila)
                
                # ===== PROCESAR CÓDIGOS N1 Y N2 =====
                def agregar_codigos(tipo_codigo, codigos_data):
                    # Diccionario para agrupar códigos por clave (potencia)
                    codigos_agrupados = {}

                    for key, nodos_data in codigos_data.items():
                        # Si esta clave no está en el diccionario, inicializarla
                        if key not in codigos_agrupados:
                            codigos_agrupados[key] = {
                                'total': 0,
                                'por_nodo': {nodo: set() for nodo in nodos_ordenados}
                            }

                        # Contar códigos por nodo y total
                        for nodo, codigos in nodos_data.items():
                            codigos_agrupados[key]['total'] += len(codigos)
                            if nodo in nodos_ordenados:
                                codigos_agrupados[key]['por_nodo'][nodo].update(codigos)

                    # Ahora, crear las filas con los códigos agrupados
                    for key, datos in codigos_agrupados.items():
                        total = datos['total']
                        # Crear la fila con suficientes elementos vacíos para todos los nodos
                        fila = [key, 'UND', total, ''] + [''] * len(nodos_ordenados)

                        # Agregar códigos por nodo
                        for i, nodo in enumerate(nodos_ordenados):
                            # Obtener los códigos para este nodo
                            codigos = datos['por_nodo'][nodo]
                            # Si hay códigos, mostrarlos separados por coma
                            if codigos:
                                fila[4 + i] = ', '.join(sorted(codigos))

                        filas.append(fila)

                if 'codigos_n1' in info:
                    agregar_codigos("N1", info['codigos_n1'])
                if 'codigos_n2' in info:
                    agregar_codigos("N2", info['codigos_n2'])
                
                # ===== MATERIALES INSTALADOS =====
                filas.append(['MATERIALES INSTALADOS', '', '', ''] + [''] * num_nodos)
                for material_key in info['materiales']:
                    try:
                        _, nombre = material_key.split('|', 1)
                        cantidades = info['materiales'][material_key]
                        total = sum(cantidades.values())
                        # Celda de fecha vacía para materiales instalados
                        fila = [nombre, 'UND', total, ''] + [cantidades.get(n, 0) for n in nodos_ordenados]
                        filas.append(fila)
                    except Exception as e:
                        logger.error(f"Error procesando material: {str(e)}")
                        continue                                
                
                # ===== MATERIALES RETIRADOS =====
                filas.append(['MATERIALES RETIRADOS', '', '', ''] + [''] * num_nodos)
                for material_key in info.get('materiales_retirados', {}):
                    try:
                        _, nombre = material_key.split('|', 1)
                        cantidades = info['materiales_retirados'][material_key]
                        total = sum(cantidades.values())
                        # Celda de fecha vacía para materiales retirados
                        fila = [nombre, 'UND', total, ''] + [cantidades.get(n, 0) for n in nodos_ordenados]
                        filas.append(fila)
                    except Exception as e:
                        logger.error(f"Error procesando material retirado: {str(e)}")
                        continue
                
                # ===== OBSERVACIONES COMPLETAS =====
                filas.append(['OBSERVACIONES COMPLETAS', '', '', ''] + [''] * num_nodos)
                
                # Organizar observaciones por nodo para la sección "OBSERVACIONES COMPLETAS"
                observaciones_completas = {nodo: [] for nodo in nodos_ordenados}
                
                # Recopilar todas las observaciones con sus códigos y fechas
                for nodo in nodos_ordenados:
                    poste = nodo.split('_')[0]
                    fecha = info['fechas_sync'].get(nodo, "Sin fecha")
                    
                    # Obtener códigos asociados a este nodo específico
                    codigos = set()
                    for key in info['codigos_n1']:
                        if nodo in info['codigos_n1'][key]:
                            codigos.update(info['codigos_n1'][key].get(nodo, set()))
                    for key in info['codigos_n2']:
                        if nodo in info['codigos_n2'][key]:
                            codigos.update(info['codigos_n2'][key].get(nodo, set()))
                    
                    if not codigos:
                        codigos.add("Sin código")
                    
                    # Obtener observaciones específicas para este nodo
                    aspectos = set()
                    for mat_key in info['aspectos_materiales']:
                        if nodo in info['aspectos_materiales'][mat_key]:
                            aspectos.update(info['aspectos_materiales'][mat_key].get(nodo, []))
                    for mat_key in info['aspectos_retirados']:
                        if nodo in info['aspectos_retirados'][mat_key]:
                            aspectos.update(info['aspectos_retirados'][mat_key].get(nodo, []))
                    
                    if not aspectos:
                        aspectos.add("Sin observaciones")
                    
                    # Formatear y agregar a las observaciones completas
                    for codigo in sorted(codigos):
                        entrada = f"Código: {codigo}\n"
                        entrada += '\n'.join([f"{i+1}. {obs}" for i, obs in enumerate(sorted(aspectos))])
                        # Agregar la fecha al final de la entrada
                        entrada += f"\n(Fecha: {fecha})"
                        observaciones_completas[nodo].append(entrada)
                
                # Crear una única fila para todas las observaciones completas
                fila_obs_completas = ['Observaciones Completas', 'Obs', '', ''] + [''] * num_nodos
                
                # Para cada nodo, formatear sus observaciones completas y colocarlas en la columna correspondiente
                for i, nodo in enumerate(nodos_ordenados):
                    if observaciones_completas[nodo]:
                        # Separar cada entrada con una línea en blanco
                        texto_obs = '\n\n'.join(observaciones_completas[nodo])
                        fila_obs_completas[4 + i] = texto_obs
                
                # Agregar la fila de observaciones completas
                filas.append(fila_obs_completas)
                
                # ===== MANO DE OBRA POR NODO =====
                # Agregar encabezado para la mano de obra
                filas.append(['MANO DE OBRA POR NODO', '', '', ''] + [''] * num_nodos)
                
                # Procesar materiales instalados y retirados para calcular la mano de obra
                materiales_instalados = {}
                materiales_retirados = {}
                
                # Extraer materiales instalados
                for material_key, nodos_qty in info['materiales'].items():
                    materiales_instalados[material_key] = {}
                    for nodo, qty in nodos_qty.items():
                        if nodo in nodos_ordenados and qty > 0:
                            materiales_instalados[material_key][nodo] = float(qty)
                
                # Extraer materiales retirados
                for material_key, nodos_qty in info.get('materiales_retirados', {}).items():
                    materiales_retirados[material_key] = {}
                    for nodo, qty in nodos_qty.items():
                        if nodo in nodos_ordenados and qty > 0:
                            materiales_retirados[material_key][nodo] = float(qty)
                
                # Cargar la plantilla de mano de obra
                plantilla = cargar_plantilla_mano_obra()
                
                # Agrupar partidas por bloques para mejor visualización
                bloques = [
                    ("Postes",
                    ["APERTURA", "APLOMADA", "CONCRETADA", "HINCADA", "ALQUILER"],  # Removido "BOTADO DE ESCOMBROS"
                    ["POSTE"]),
                    ("Instalación luminarias",
                    ["INSTALACION LUMINARIAS"],
                    ["LUMINARIA", "FOTOCELDA", "GRILLETE", "BRAZO"]),
                    ("Conexión a tierra",
                    ["CONEXIÓN A CABLE A TIERRA", "INSTALACION KIT SPT", "INSTALACION DE ATERRIZAJES", "BOTADO DE ESCOMBROS", "RECUPERACION ZONA DURA"],  # Específico para zona dura
                    ["KIT DE PUESTA A TIERRA", "CONECT PERF", "CONECTOR BIME/COM", "ALAMBRE", "TUERCA", "TORNILLO", "VARILLA"]),
                    ("Desmontaje / Transporte",
                    ["DESMONTAJE", "TRANSPORTE", "TRANSP.", "TRANSPORTE COLLARINES"],
                    ["ALAMBRE", "BRAZO", "CÓDIGO", "CABLE", "ABRAZADERA", "GRILLETE"]),
                    ("Instalación de cables",
                    ["INSTALACION CABLE"],
                    ["CABLE", "TPX", "ALAMBRE"]),
                    ("Otros trabajos",
                    ["VESTIDA CONJUNTO 1 O 2 PERCHAS", "VESTIDA CONJUNTO 3 O MAS PERCHAS", "VESTIDA", "CAJA", "PINTADA", "EXCAVACION", "RECUPERACION ZONA", "SOLDADURA", "INSTALACION TRAMA", "INSTALACION CORAZA"],  # Solo "RECUPERACION ZONA" sin "DURA"
                    ["PERCHA", "CAJA", "TUBO", "TUBERIA", "CONDUIT"])
                ]
                
                # Pre-procesamiento para unificar los tipos de desmontaje de luminarias
                partidas_unificadas = []
                tipo_desmontaje_luminarias = []
                
                #for partida in plantilla:
                #    descripcion = partida['DESCRIPCION MANO DE OBRA']
                #    # Verificar si es una partida de desmontaje de luminarias
                #    if ("DESMONTAJE" in descripcion.upper() and "LUMINARIA" in descripcion.upper() and 
                #        ("CAMIONETA" in descripcion.upper() or "CANASTA" in descripcion.upper())):
                #        # Guardar la descripción original para referencia
                #        tipo_desmontaje_luminarias.append(descripcion)
                #        # No agregar esta partida a la lista unificada aún
                #    else:
                #        partidas_unificadas.append(partida)
                #
                ## Si hay partidas de desmontaje de luminarias, crear una partida unificada
                #if tipo_desmontaje_luminarias:
                #    partida_unificada = {
                #        'DESCRIPCION MANO DE OBRA': "DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA",
                #        'UNIDAD': next((p['UNIDAD'] for p in plantilla if p['DESCRIPCION MANO DE OBRA'] in tipo_desmontaje_luminarias), "UN")
                #    }
                #    partidas_unificadas.append(partida_unificada)
                #
                ## Reemplazar la plantilla original con la unificada
                #plantilla_original = plantilla
                #plantilla = partidas_unificadas
                plantilla_original = plantilla
                # Crear un diccionario para almacenar la mano de obra por nodo
                mano_obra_por_nodo = {nodo: {} for nodo in nodos_ordenados}

                # Crear un diccionario para almacenar información sobre el tipo de brazo por nodo
                info_brazos_por_nodo = {}
                
                # MEJORA: Determinar si hay brazos grandes o pequeños en cada nodo
                for nodo in nodos_ordenados:
                    # Verificar si hay LUMINARIA CODIGO/BRAZO en este nodo específico
                    tiene_luminaria_codigo_brazo = False
                    cantidad_luminaria_codigo_brazo = 0
                    
                    for material_key, nodos_qty in materiales_instalados.items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                tiene_luminaria_codigo_brazo = True
                                cantidad_luminaria_codigo_brazo += nodos_qty[nodo]
                    
                    # Solo verificar brazos si hay LUMINARIA CODIGO/BRAZO
                    usar_canasta = False
                    brazo_encontrado = False
                    tamano_brazo = 0
                    
                    if tiene_luminaria_codigo_brazo:
                        # Verificar si hay brazos INSTALADOS en este nodo específico y determinar su tamaño
                        for material_key, nodos_qty in materiales_instalados.items():
                            if "|" in material_key:
                                material_name = material_key.split("|")[1].upper()
                                if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                    brazo_encontrado = True
                                    # Intentar extraer la longitud del brazo
                                    longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                                    if longitud_match:
                                        try:
                                            longitud = int(longitud_match.group(1))
                                            tamano_brazo = longitud
                                            # Considerar brazos de 3 metros como grandes (para canasta)
                                            if longitud >= 3:  # Si el brazo es mayor o igual a 3 metros, usar canasta
                                                usar_canasta = True
                                                break
                                        except:
                                            pass
                                    # Si no se puede extraer la longitud pero contiene indicación de tamaño grande
                                    elif "2 1/2" in material_name or "2.5" in material_name:
                                        usar_canasta = True
                                        tamano_brazo = 4  # Asumimos un valor mayor a 3
                                        break
                                    # Verificar explícitamente si es un brazo de 3 metros
                                    elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                                        usar_canasta = True
                                        tamano_brazo = 3
                                        break
                                    
                    # Guardar la información del tipo de brazo para este nodo
                    info_brazos_por_nodo[nodo] = {
                        'brazo_encontrado': brazo_encontrado,
                        'tamano_brazo': tamano_brazo,
                        'usar_canasta': usar_canasta,
                        'tiene_luminaria_codigo_brazo': tiene_luminaria_codigo_brazo,
                        'cantidad_luminaria_codigo_brazo': cantidad_luminaria_codigo_brazo
                    }
                    
                    # Inicializar el bloque de instalación de luminarias si no existe
                    if "Instalación luminarias" not in mano_obra_por_nodo[nodo]:
                        mano_obra_por_nodo[nodo]["Instalación luminarias"] = []
                    
                    # Solo agregar partida si hay LUMINARIA CODIGO/BRAZO
                    if tiene_luminaria_codigo_brazo:
                        # Elegir la descripción adecuada según el tipo de brazo
                        desc_instalacion = "INSTALACION DE LUMINARIAS EN CANASTA" if usar_canasta else "INSTALACION DE LUMINARIAS EN CAMIONETA"
                        
                        # Verificar si ya existe esta partida
                        ya_existe = False
                        for partida in mano_obra_por_nodo[nodo].get("Instalación luminarias", []):
                            if partida['descripcion'] == desc_instalacion:
                                ya_existe = True
                                break
                            
                        # Solo agregar si no existe
                        if not ya_existe:
                            # Preparar información sobre el brazo
                            info_brazo = ""
                            if brazo_encontrado:
                                if usar_canasta:
                                    info_brazo = f"BRAZO GRANDE ({tamano_brazo}M)"
                                else:
                                    info_brazo = f"BRAZO PEQUEÑO ({tamano_brazo}M)"
                            
                            # Combinar información de luminarias y brazos
                            materiales = []
                            materiales.append(f"LUMINARIA CODIGO/BRAZO: {cantidad_luminaria_codigo_brazo} unidades")
                            if info_brazo:
                                materiales.append(info_brazo)
                            
                            # Agregar la partida
                            mano_obra_por_nodo[nodo]["Instalación luminarias"].append({
                                'descripcion': desc_instalacion,
                                'unidad': "UND",
                                'cantidad': cantidad_luminaria_codigo_brazo,  # Usar la cantidad exacta de LUMINARIA CODIGO/BRAZO
                                'materiales': ""
                            })
                
                info_desmontaje_por_nodo = {}

                for nodo in nodos_ordenados:
                    # Verificar si hay brazos RETIRADOS en este nodo específico y determinar su tamaño
                    usar_canasta_desmontaje = False
                    brazo_retirado_encontrado = False
                    tamano_brazo_retirado = 0
                    luminarias_retiradas = 0
                
                    # Contar luminarias retiradas
                    for material_key, nodos_qty in materiales_retirados.items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            # Verificar si es una luminaria retirada
                            es_luminaria_retirada = (
                                "LUMINARIA" in material_name or 
                                "LUM" in material_name or
                                "LED" in material_name or
                                "LAMP" in material_name or
                                "FOCO" in material_name or
                                "RETIRADA" in material_name
                            )
                            if es_luminaria_retirada and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                luminarias_retiradas += nodos_qty[nodo]
                
                    # Verificar brazos retirados
                    for material_key, nodos_qty in materiales_retirados.items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                brazo_retirado_encontrado = True
                                # Intentar extraer la longitud del brazo
                                longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                                if longitud_match:
                                    try:
                                        longitud = int(longitud_match.group(1))
                                        tamano_brazo_retirado = longitud
                                        # Considerar brazos de 3 metros como grandes (para canasta)
                                        if longitud >= 3:  # Si el brazo es mayor o igual a 3 metros, usar canasta
                                            usar_canasta_desmontaje = True
                                            break
                                    except:
                                        pass
                                # Si no se puede extraer la longitud pero contiene indicación de tamaño grande
                                elif "2 1/2" in material_name or "2.5" in material_name:
                                    usar_canasta_desmontaje = True
                                    tamano_brazo_retirado = 4  # Asumimos un valor mayor a 3
                                    break
                                # Verificar explícitamente si es un brazo de 3 metros
                                elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                                    usar_canasta_desmontaje = True
                                    tamano_brazo_retirado = 3
                                    break
                                
                    # Guardar la información del desmontaje para este nodo
                    info_desmontaje_por_nodo[nodo] = {
                        'brazo_retirado_encontrado': brazo_retirado_encontrado,
                        'tamano_brazo_retirado': tamano_brazo_retirado,
                        'usar_canasta_desmontaje': usar_canasta_desmontaje,
                        'luminarias_retiradas': luminarias_retiradas
                    }
                    
                    # Inicializar el bloque de desmontaje de luminarias si no existe
                    if "Desmontaje / Transporte" not in mano_obra_por_nodo[nodo]:
                        mano_obra_por_nodo[nodo]["Desmontaje / Transporte"] = []
                    
                    # Si hay luminarias retiradas, agregar la partida correspondiente
                    if luminarias_retiradas > 0:
                        # Elegir la descripción adecuada según el tipo de brazo
                        desc_desmontaje = "DESMONTAJE DE LUMINARIAS EN CANASTA" if usar_canasta_desmontaje else "DESMONTAJE DE LUMINARIAS EN CAMIONETA"
                        
                        # MODIFICACIÓN: Separar desmontajes por tipo (CANASTA vs CAMIONETA)
                # Primero, recopilar información sobre brazos retirados y luminarias por nodo
                brazos_grandes_retirados_por_nodo = {}
                brazos_pequenos_retirados_por_nodo = {}
                luminarias_retiradas_por_nodo = {}
                
                
                # Recopilar información sobre brazos retirados por nodo
                for nodo in nodos_ordenados:
                    brazos_grandes_retirados_por_nodo[nodo] = []
                    brazos_pequenos_retirados_por_nodo[nodo] = []
                    luminarias_retiradas_por_nodo[nodo] = []
                
                    # Buscar luminarias retiradas en este nodo
                    for material_key, nodos_qty in materiales_retirados.items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            # Verificar si es una luminaria retirada
                            es_luminaria_retirada = (
                                "LUMINARIA" in material_name or 
                                "LUM" in material_name or
                                "LED" in material_name or
                                "LAMP" in material_name or
                                "FOCO" in material_name or
                                "RETIRADA" in material_name
                            )
                            if es_luminaria_retirada and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                luminarias_retiradas_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                
                    # Buscar brazos retirados en materiales retirados y clasificarlos por tamaño
                    for material_key, nodos_qty in materiales_retirados.items():
                        if "|" in material_key:
                            material_name = material_key.split("|")[1].upper()
                            if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                # Intentar extraer la longitud del brazo
                                longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                                if longitud_match:
                                    try:
                                        longitud = int(longitud_match.group(1))
                                        if longitud >= 3:  # Si el brazo es mayor a 3 metros, usar canasta
                                            brazos_grandes_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                        else:
                                            brazos_pequenos_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                    except:
                                        # Si no se puede extraer la longitud pero contiene indicación de tamaño
                                        if "2 1/2" in material_name or "2.5" in material_name:
                                            brazos_grandes_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                        else:
                                            brazos_pequenos_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                # Si no se puede extraer la longitud pero contiene indicación de tamaño grande
                                elif "2 1/2" in material_name or "2.5" in material_name:
                                    brazos_grandes_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                # Verificar explícitamente si es un brazo de 3 metros
                                elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                                    brazos_grandes_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                else:
                                    brazos_pequenos_retirados_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                
                # Crear fila para DESMONTAJE DE LUMINARIAS EN CANASTA
                total_canasta_desmontaje = 0
                luminarias_canasta_desmontaje = set()
                brazos_canasta_desmontaje = set()
                
                # Crear fila para DESMONTAJE DE LUMINARIAS EN CAMIONETA
                total_camioneta_desmontaje = 0
                luminarias_camioneta_desmontaje = set()
                brazos_camioneta_desmontaje = set()
                
                # Procesar cada nodo para determinar qué tipo de desmontaje necesita
                for nodo in nodos_ordenados:
                    # Si hay brazos grandes retirados, asignar a CANASTA
                    if brazos_grandes_retirados_por_nodo[nodo]:
                        # La cantidad es el número de luminarias retiradas o al menos 1 si hay brazos grandes
                        cantidad_luminarias = len(luminarias_retiradas_por_nodo[nodo])
                        cantidad_nodo = max(cantidad_luminarias, 1) if brazos_grandes_retirados_por_nodo[nodo] else 0
                        
                        if cantidad_nodo > 0:
                            total_canasta_desmontaje += cantidad_nodo
                
                            # Agregar luminarias a la lista para CANASTA
                            for luminaria in luminarias_retiradas_por_nodo[nodo]:
                                luminarias_canasta_desmontaje.add(luminaria)
                
                            # Agregar brazos a la lista de brazos para CANASTA
                            for brazo in brazos_grandes_retirados_por_nodo[nodo]:
                                brazos_canasta_desmontaje.add(brazo)
                
                    # Si hay brazos pequeños o luminarias sin brazos grandes, asignar a CAMIONETA
                    elif brazos_pequenos_retirados_por_nodo[nodo] or luminarias_retiradas_por_nodo[nodo]:
                        # La cantidad es el número de luminarias retiradas o al menos 1 si hay brazos pequeños
                        cantidad_luminarias = len(luminarias_retiradas_por_nodo[nodo])
                        cantidad_nodo = max(cantidad_luminarias, 1) if brazos_pequenos_retirados_por_nodo[nodo] or luminarias_retiradas_por_nodo[nodo] else 0
                        
                        if cantidad_nodo > 0:
                            total_camioneta_desmontaje += cantidad_nodo
                
                            # Agregar luminarias a la lista para CAMIONETA
                            for luminaria in luminarias_retiradas_por_nodo[nodo]:
                                luminarias_camioneta_desmontaje.add(luminaria)
                
                            # Agregar brazos a la lista de brazos para CAMIONETA
                            for brazo in brazos_pequenos_retirados_por_nodo[nodo]:
                                brazos_camioneta_desmontaje.add(brazo)
                
                # Crear fila para DESMONTAJE DE LUMINARIAS EN CANASTA si hay cantidad
                #if total_canasta_desmontaje > 0:
                #    descripcion = "DESMONTAJE DE LUMINARIAS EN CANASTA"
                #    unidad = "UND"
                #
                #    # Crear resumen de materiales
                #    todos_materiales = []
                #    if luminarias_canasta_desmontaje:
                #        todos_materiales.append(f"RET: LUMINARIAS: {', '.join(sorted(luminarias_canasta_desmontaje))}")
                #    if brazos_canasta_desmontaje:
                #        todos_materiales.append(f"RET: BRAZOS: {', '.join(sorted(brazos_canasta_desmontaje))}")
                #
                #    # Crear la fila
                #    fila_canasta_desmontaje = [descripcion, unidad, total_canasta_desmontaje, ''] + [''] * num_nodos
                #
                #    # Agregar materiales en la columna de fecha
                #    if todos_materiales:
                #        fila_canasta_desmontaje[3] = "\n".join(todos_materiales)
                #
                #    # Para cada nodo, agregar la cantidad correspondiente
                #    for i, nodo in enumerate(nodos_ordenados):
                #        if brazos_grandes_retirados_por_nodo[nodo]:
                #            # La cantidad es el número de luminarias o al menos 1 si hay brazos grandes
                #            cantidad_luminarias = len(luminarias_retiradas_por_nodo[nodo])
                #            cantidad_nodo = max(cantidad_luminarias, 1)
                #
                #            # Crear texto con materiales para este nodo
                #            materiales_nodo = []
                #            if luminarias_retiradas_por_nodo[nodo]:
                #                materiales_nodo.append(f"RET: LUMINARIAS: {', '.join(luminarias_retiradas_por_nodo[nodo])}")
                #            if brazos_grandes_retirados_por_nodo[nodo]:
                #                materiales_nodo.append(f"RET: BRAZOS: {', '.join(brazos_grandes_retirados_por_nodo[nodo])}")
                #
                #            # Formatear contenido: cantidad + materiales
                #            contenido = f"{cantidad_nodo}"
                #            if materiales_nodo:
                #                contenido += f"\n{chr(10).join(materiales_nodo)}"
                #
                #            fila_canasta_desmontaje[4 + i] = contenido
                #
                #    # Agregar la fila a la lista de filas
                #    filas.append(fila_canasta_desmontaje)
                
                # Crear fila para DESMONTAJE DE LUMINARIAS EN CAMIONETA si hay cantidad
                #if total_camioneta_desmontaje > 0:
                #    descripcion = "DESMONTAJE DE LUMINARIAS EN CAMIONETA"
                #    unidad = "UND"
                #
                #    # Crear resumen de materiales
                #    todos_materiales = []
                #    if luminarias_camioneta_desmontaje:
                #        todos_materiales.append(f"RET: LUMINARIAS: {', '.join(sorted(luminarias_camioneta_desmontaje))}")
                #    if brazos_camioneta_desmontaje:
                #        todos_materiales.append(f"RET: BRAZOS: {', '.join(sorted(brazos_camioneta_desmontaje))}")
                #
                #    # Crear la fila
                #    fila_camioneta_desmontaje = [descripcion, unidad, total_camioneta_desmontaje, ''] + [''] * num_nodos
                #
                #    # Agregar materiales en la columna de fecha
                #    if todos_materiales:
                #        fila_camioneta_desmontaje[3] = "\n".join(todos_materiales)
                #
                #    # Para cada nodo, agregar la cantidad correspondiente
                #    for i, nodo in enumerate(nodos_ordenados):
                #        if (brazos_pequenos_retirados_por_nodo[nodo] or luminarias_retiradas_por_nodo[nodo]) and not brazos_grandes_retirados_por_nodo[nodo]:
                #            # La cantidad es el número de luminarias o al menos 1 si hay brazos pequeños
                #            cantidad_luminarias = len(luminarias_retiradas_por_nodo[nodo])
                #            cantidad_nodo = max(cantidad_luminarias, 1) if brazos_pequenos_retirados_por_nodo[nodo] or luminarias_retiradas_por_nodo[nodo] else 0
                #
                #            if cantidad_nodo > 0:
                #                # Crear texto con materiales para este nodo
                #                materiales_nodo = []
                #                if luminarias_retiradas_por_nodo[nodo]:
                #                    materiales_nodo.append(f"RET: LUMINARIAS: {', '.join(luminarias_retiradas_por_nodo[nodo])}")
                #                if brazos_pequenos_retirados_por_nodo[nodo]:
                #                    materiales_nodo.append(f"RET: BRAZOS: {', '.join(brazos_pequenos_retirados_por_nodo[nodo])}")
                #
                #                # Formatear contenido: cantidad + materiales
                #                contenido = f"{cantidad_nodo}"
                #                if materiales_nodo:
                #                    contenido += f"\n{chr(10).join(materiales_nodo)}"
                #
                #                fila_camioneta_desmontaje[4 + i] = contenido
                #
                #    # Agregar la fila a la lista de filas
                #    filas.append(fila_camioneta_desmontaje)
                
                # Para cada nodo, calcular la mano de obra necesaria
                for nodo in nodos_ordenados:
                    # Para cada bloque de partidas
                    for titulo_bloque, keywords_mo, keywords_mat in bloques:
                        # Si es el bloque de instalación de luminarias, ya lo procesamos antes
                        if titulo_bloque == "Instalación luminarias":
                            continue
                            
                        # Filtrar partidas del bloque actual
                        partidas_filtradas = [
                            item for item in plantilla
                            if any(kw in item['DESCRIPCION MANO DE OBRA'].upper() for kw in keywords_mo)
                        ]
                        
                        # Para cada partida, calcular la mano de obra necesaria para este nodo
                        for partida in partidas_filtradas:
                            descripcion = partida['DESCRIPCION MANO DE OBRA']
                            unidad = partida['UNIDAD']
                            
                            # Caso especial para DESMONTAJE DE LUMINARIAS CANASTA/ESCALERA
                            cantidad_mo, materiales_inst, materiales_ret = calcular_cantidad_mano_obra(
                                descripcion,
                                materiales_instalados,
                                materiales_retirados,
                                nodo,
                                info.get('codigos_n1', {}),
                                info.get('codigos_n2', {}),
                                titulo_bloque,
                                
                            )
                            
                            # Solo agregar partidas con cantidad > 0
                            if cantidad_mo > 0:
                                # Si este bloque no existe en el diccionario del nodo, crearlo
                                if titulo_bloque not in mano_obra_por_nodo[nodo]:
                                    mano_obra_por_nodo[nodo][titulo_bloque] = []
                                
                                # Formatear materiales asociados
                                materiales_texto = []
                                #if materiales_inst:
                                #    materiales_texto.append("INST: " + ", ".join(materiales_inst))
                                #if materiales_ret:
                                #    materiales_texto.append("RET: " + ", ".join(materiales_ret))
                                
                                # Agregar la partida al bloque correspondiente
                                mano_obra_por_nodo[nodo][titulo_bloque].append({
                                    'descripcion': descripcion,
                                    'unidad': unidad,
                                    'cantidad': cantidad_mo,
                                    'materiales': ""
                                    #'materiales': "\n".join(materiales_texto)
                                })
                
                # IMPORTANTE: Forzar la aparición de instalación de luminarias en al menos un nodo
                # si hay códigos N1 o N2 en la OT pero no se calculó mano de obra
                hay_codigos_ot = False
                for _, nodos_valores in info.get('codigos_n1', {}).items():
                    for nodo, codigos in nodos_valores.items():
                        # Filtrar códigos NO, N/A, NA
                        codigos_validos = [c for c in codigos if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                        if codigos_validos:
                            hay_codigos_ot = True
                            break
                    if hay_codigos_ot:
                        break
                    
                if not hay_codigos_ot:
                    for _, nodos_valores in info.get('codigos_n2', {}).items():
                        for nodo, codigos in nodos_valores.items():
                            # Filtrar códigos NO, N/A, NA
                            codigos_validos = [c for c in codigos if c.upper() not in ["NO", "N/A", "NA", "NO APLICA"]]
                            if codigos_validos:
                                hay_codigos_ot = True
                                break
                        if hay_codigos_ot:
                            break
                
                # Verificar si ya hay alguna partida de instalación de luminarias
                hay_instalacion_luminarias = False
                for nodo in nodos_ordenados:
                    if "Instalación luminarias" in mano_obra_por_nodo[nodo]:
                        for partida in mano_obra_por_nodo[nodo]["Instalación luminarias"]:
                            if "INSTALACION DE LUMINARIAS" in partida['descripcion'].upper():
                                hay_instalacion_luminarias = True
                                break
                        if hay_instalacion_luminarias:
                            break
                        
                # Verificar si hay códigos válidos en la OT (excluyendo NO, N/A, NA)
                
                
                # FORZAR la aparición de instalación de luminarias si hay códigos pero no hay instalación
                if hay_codigos_ot and not hay_instalacion_luminarias and nodos_ordenados:
                    # Determinar si hay algún nodo con brazo grande en la OT
                    usar_canasta_este_nodo = info_brazos_por_nodo.get(nodo_con_codigos, {}).get('usar_canasta', False)
                    
                    # Elegir el primer nodo que tenga códigos
                    nodo_con_codigos = None
                    for nodo in nodos_ordenados:
                        tiene_codigos = False
                        for _, nodos_valores in info.get('codigos_n1', {}).items():
                            if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                                tiene_codigos = True
                                break
                        
                        if not tiene_codigos:
                            for _, nodos_valores in info.get('codigos_n2', {}).items():
                                if nodo in nodos_valores and len(nodos_valores[nodo]) > 0:
                                    tiene_codigos = True
                                    break
                        
                        if tiene_codigos:
                            nodo_con_codigos = nodo
                            break
                    
                    # Si no hay nodo con códigos, usar el primer nodo
                    if not nodo_con_codigos and nodos_ordenados:
                        nodo_con_codigos = nodos_ordenados[0]
                    
                    if nodo_con_codigos:
                        # Inicializar el bloque si no existe
                        if "Instalación luminarias" not in mano_obra_por_nodo[nodo_con_codigos]:
                            mano_obra_por_nodo[nodo_con_codigos]["Instalación luminarias"] = []
                        
                        # Determinar si usar canasta basado en la presencia de brazos grandes en la OT
                        desc_instalacion = "INSTALACION DE LUMINARIAS EN CANASTA" if usar_canasta_este_nodo else "INSTALACION DE LUMINARIAS EN CAMIONETA"
                        
                        # Obtener los códigos para mostrar
                        codigos_texto = []
                        for key_codigo, nodos_valores in info.get('codigos_n1', {}).items():
                            for nodo, codigos in nodos_valores.items():
                                if codigos:
                                    codigos_texto.append(f"CÓDIGO N1: {', '.join(codigos)}")
                        
                        for key_codigo, nodos_valores in info.get('codigos_n2', {}).items():
                            for nodo, codigos in nodos_valores.items():
                                if codigos:
                                    codigos_texto.append(f"CÓDIGO N2: {', '.join(codigos)}")
                        
                        # Formatear materiales
                        #materiales = "INST: " + (", ".join(codigos_texto) if codigos_texto else "LUMINARIA INSTALADA (FORZADO)")
                        
                        # Agregar la partida forzada
                        mano_obra_por_nodo[nodo_con_codigos]["Instalación luminarias"].append({
                            'descripcion': desc_instalacion,
                            'unidad': "UND",
                            'cantidad': 1,
                            'materilaes': ""
                            #'materiales': materiales
                        })
                
                # Crear filas para cada bloque de mano de obra
                for titulo_bloque, _, _ in bloques:
                    # FORZAR la aparición del bloque de instalación de luminarias
                    if titulo_bloque == "Instalación luminarias":
                        # Siempre agregar este bloque si hay LUMINARIA CODIGO/BRAZO
                        hay_luminaria_codigo_brazo = False

                        # Verificar si hay LUMINARIA CODIGO/BRAZO en materiales instalados
                        for material_key, nodos_qty in info['materiales'].items():
                            if "|" in material_key:
                                material_name = material_key.split("|")[1].upper()
                                if "LUMINARIA CODIGO/BRAZO" in material_name:
                                    for nodo, cantidad in nodos_qty.items():
                                        if cantidad > 0:
                                            hay_luminaria_codigo_brazo = True
                                            break
                            if hay_luminaria_codigo_brazo:
                                break
                            
                        # Solo proceder si hay LUMINARIA CODIGO/BRAZO
                        if hay_luminaria_codigo_brazo:
                            filas.append([f"BLOQUE: {titulo_bloque}", '', '', ''] + [''] * num_nodos)

                            # MODIFICACIÓN: Separar instalaciones por tipo (CANASTA vs CAMIONETA)
                            # Primero, recopilar información sobre brazos y LUMINARIA CODIGO/BRAZO por nodo
                            brazos_grandes_por_nodo = {}
                            brazos_pequenos_por_nodo = {}
                            luminarias_codigo_brazo_por_nodo = {}

                            # Recopilar información sobre brazos y luminarias por nodo
                            for nodo in nodos_ordenados:
                                brazos_grandes_por_nodo[nodo] = []
                                brazos_pequenos_por_nodo[nodo] = []
                                luminarias_codigo_brazo_por_nodo[nodo] = []

                                # Contar LUMINARIA CODIGO/BRAZO para este nodo
                                for material_key, nodos_qty in info['materiales'].items():
                                    if "|" in material_key:
                                        material_name = material_key.split("|")[1].upper()
                                        if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                            luminarias_codigo_brazo_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")

                                # Buscar brazos en materiales instalados y clasificarlos por tamaño
                                for material_key, nodos_qty in materiales_instalados.items():
                                    if "|" in material_key:
                                        material_name = material_key.split("|")[1].upper()
                                        if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                            # Intentar extraer la longitud del brazo
                                            longitud_match = re.search(r'(\d+(?:\.\d+)?)\s*M', material_name)
                                            if longitud_match:
                                                try:
                                                    longitud = float(longitud_match.group(1))
                                                    # CORRECCIÓN: Solo brazos >= 3 metros son grandes
                                                    if longitud >= 3.0:
                                                        brazos_grandes_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                                    else:
                                                        brazos_pequenos_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                                except:
                                                    # Si no se puede convertir a número, clasificar como pequeño por defecto
                                                    brazos_pequenos_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                            else:
                                                # Si no hay patrón de longitud numérica, verificar patrones específicos
                                                if any(pattern in material_name for pattern in ["3 MT", "3 MTS", "3M", "4 MT", "4 MTS", "4M", "5 MT", "5 MTS", "5M", "6 MT", "6 MTS", "6M"]):
                                                    brazos_grandes_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")
                                                else:
                                                    # Por defecto, clasificar como pequeño
                                                    brazos_pequenos_por_nodo[nodo].append(f"{material_name} ({nodos_qty[nodo]})")

                            # Crear fila para INSTALACION DE LUMINARIAS EN CANASTA
                            total_canasta = 0
                            luminarias_canasta = set()
                            brazos_canasta = set()

                            # Crear fila para INSTALACION DE LUMINARIAS EN CAMIONETA
                            total_camioneta = 0
                            luminarias_camioneta = set()
                            brazos_camioneta = set()

                            # Procesar cada nodo para determinar qué tipo de instalación necesita
                            for nodo in nodos_ordenados:
                                # Contar LUMINARIA CODIGO/BRAZO en este nodo
                                cantidad_luminarias_codigo_brazo = 0
                                for material_key, nodos_qty in info['materiales'].items():
                                    if "|" in material_key:
                                        material_name = material_key.split("|")[1].upper()
                                        if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                                            cantidad_luminarias_codigo_brazo += nodos_qty[nodo]

                                # Solo proceder si hay LUMINARIA CODIGO/BRAZO en este nodo
                                if cantidad_luminarias_codigo_brazo > 0:
                                    # Si hay brazos grandes, asignar a CANASTA
                                    if brazos_grandes_por_nodo[nodo]:
                                        tiene_brazo_3m_confirmado = False
                                        for material_key, nodos_qty in materiales_instalados.items():
                                            if "|" in material_key:
                                                material_name = material_key.split("|")[1].upper()
                                                if "BRAZO" in material_name and nodo in nodos_qty and nodos_qty[nodo] > 0:
                                                    longitud_match = re.search(r'(\d+)\s*M(?:TS?)?', material_name)
                                                    if longitud_match:
                                                        try:
                                                            longitud = int(longitud_match.group(1))
                                                            if longitud >= 3:
                                                                tiene_brazo_3m_confirmado = True
                                                                break
                                                        except:
                                                            pass
                                                    # Verificar patrones específicos
                                                    elif "3 MT" in material_name or "3 MTS" in material_name or "3M" in material_name:
                                                        tiene_brazo_3m_confirmado = True
                                                        break
                                        if brazos_grandes_por_nodo[nodo]:
                                            total_canasta += cantidad_luminarias_codigo_brazo
                                            for luminaria in luminarias_codigo_brazo_por_nodo[nodo]:
                                                luminarias_canasta.add(luminaria)
                                            for brazo in brazos_grandes_por_nodo[nodo]:
                                                brazos_canasta.add(brazo)
                                        else:
                                            # Si no se confirma, asignar a camioneta
                                            total_camioneta += cantidad_luminarias_codigo_brazo
                                            # Agregar luminarias a la lista para CAMIONETA
                                            for luminaria in luminarias_codigo_brazo_por_nodo[nodo]:
                                                luminarias_camioneta.add(luminaria)
                                            # Agregar brazos a la lista de brazos para CAMIONETA
                                            for brazo in brazos_pequenos_por_nodo[nodo]:
                                                brazos_camioneta.add(brazo)

                                    # Si no hay brazos grandes, asignar a CAMIONETA
                                    else:
                                        total_camioneta += cantidad_luminarias_codigo_brazo

                                        # Agregar luminarias a la lista para CAMIONETA
                                        for luminaria in luminarias_codigo_brazo_por_nodo[nodo]:
                                            luminarias_camioneta.add(luminaria)

                                        # Agregar brazos a la lista de brazos para CAMIONETA
                                        for brazo in brazos_pequenos_por_nodo[nodo]:
                                            brazos_camioneta.add(brazo)

                            # Crear fila para INSTALACION DE LUMINARIAS EN CANASTA si hay cantidad
                            if total_canasta > 0:
                                descripcion = "INSTALACION DE LUMINARIAS EN CANASTA"
                                unidad = "UND"

                                # Crear la fila
                                fila_canasta = [descripcion, unidad, total_canasta, ''] + [''] * num_nodos

                                # Para cada nodo, agregar la cantidad correspondiente
                                for i, nodo in enumerate(nodos_ordenados):
                                    # Verificar si hay LUMINARIA CODIGO/BRAZO en este nodo
                                    cantidad_nodo = 0
                                    for material_key, nodos_qty in info['materiales'].items():
                                        if "|" in material_key:
                                            material_name = material_key.split("|")[1].upper()
                                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                                                cantidad_nodo += nodos_qty[nodo]

                                    # Solo asignar si hay LUMINARIA CODIGO/BRAZO y brazos grandes
                                    if cantidad_nodo > 0 and brazos_grandes_por_nodo[nodo]:
                                        fila_canasta[4 + i] = cantidad_nodo

                                # Agregar la fila a la lista de filas
                                filas.append(fila_canasta)

                            # Crear fila para INSTALACION DE LUMINARIAS EN CAMIONETA
                            if total_camioneta > 0:
                                descripcion = "INSTALACION DE LUMINARIAS EN CAMIONETA"
                                unidad = "UND"

                                # Crear la fila
                                fila_camioneta = [descripcion, unidad, total_camioneta, ''] + [''] * num_nodos

                                # Para cada nodo, agregar la cantidad correspondiente
                                for i, nodo in enumerate(nodos_ordenados):
                                    # Verificar si hay LUMINARIA CODIGO/BRAZO en este nodo
                                    cantidad_nodo = 0
                                    for material_key, nodos_qty in info['materiales'].items():
                                        if "|" in material_key:
                                            material_name = material_key.split("|")[1].upper()
                                            if "LUMINARIA CODIGO/BRAZO" in material_name and nodo in nodos_qty:
                                                cantidad_nodo += nodos_qty[nodo]

                                    # Solo asignar si hay LUMINARIA CODIGO/BRAZO y no hay brazos grandes
                                    if cantidad_nodo > 0 and not brazos_grandes_por_nodo[nodo]:
                                        fila_camioneta[4 + i] = cantidad_nodo

                                # Agregar la fila a la lista de filas
                                filas.append(fila_camioneta)

                        total_horizontal = total_canasta + total_camioneta
                        
                        #if total_horizontal > 0:
                        #    descripcion = "INSTALACION DE LUMINARIAS HORIZONTAL ADOSADA"
                        #    unidad = "UND"
                        #    
                        #    # Crear resumen de materiales combinando los de canasta y camioneta
                        #    todos_materiales = []
                        #    todos_codigos = set()
                        #    todos_codigos.update(codigos_canasta)
                        #    todos_codigos.update(codigos_camioneta)
                        #    
                        #    #if todos_codigos:
                        #    #    todos_materiales.append(f"INST: CÓDIGO N1/N2: {', '.join(sorted(todos_codigos))}")
                        #    
                        #    # Crear la fila
                        #    fila_horizontal = [descripcion, unidad, total_horizontal, ''] + [''] * num_nodos
                        #    
                        #    # Agregar materiales en la columna de fecha
                        #    if todos_materiales:
                        #        fila_horizontal[3] = "\n".join(todos_materiales)
                        #    
                        #    # Para cada nodo, agregar la cantidad correspondiente
                        #    for i, nodo in enumerate(nodos_ordenados):
                        #        cantidad_nodo = 0
                        #        
                        #        # Sumar las cantidades de canasta y camioneta para este nodo
                        #        if brazos_grandes_por_nodo[nodo]:
                        #            cantidad_nodo += max(len(codigos_por_nodo[nodo]), 1)
                        #        elif (brazos_pequenos_por_nodo[nodo] or codigos_por_nodo[nodo]):
                        #            cantidad_nodo += max(len(codigos_por_nodo[nodo]), 1)
                        #        
                        #        if cantidad_nodo > 0:
                        #            # Formatear contenido: cantidad + materiales
                        #            contenido = f"{cantidad_nodo}"
                        #            
                        #            # Crear texto con materiales para este nodo
                        #            materiales_nodo = []
                        #            #if codigos_por_nodo[nodo]:
                        #            #    materiales_nodo.append(f"INST: CÓDIGO N1/N2: {', '.join(codigos_por_nodo[nodo])}")
                        #            
                        #            if materiales_nodo:
                        #                contenido += f"\n{chr(10).join(materiales_nodo)}"
                        #            
                        #            fila_horizontal[4 + i] = contenido
                        #    
                        #    # Agregar la fila a la lista de filas
                        #    filas.append(fila_horizontal)
                        # Continuar con el resto de partidas de este bloque (que no sean instalación de luminarias)
                        for descripcion in sorted(set(
                            partida['descripcion'] 
                            for nodo in nodos_ordenados 
                            if titulo_bloque in mano_obra_por_nodo[nodo] 
                            for partida in mano_obra_por_nodo[nodo][titulo_bloque]
                            if "INSTALACION DE LUMINARIAS" not in partida['descripcion'].upper()
                        )):
                            # Procesar otras partidas del bloque que no sean instalación de luminarias
                            unidad = next((
                                partida['unidad'] 
                                for nodo in nodos_ordenados 
                                if titulo_bloque in mano_obra_por_nodo[nodo] 
                                for partida in mano_obra_por_nodo[nodo][titulo_bloque] 
                                if partida['descripcion'] == descripcion
                            ), "UND")

                            cantidad_total = sum(
                                partida['cantidad']
                                for nodo in nodos_ordenados
                                if titulo_bloque in mano_obra_por_nodo[nodo]
                                for partida in mano_obra_por_nodo[nodo][titulo_bloque]
                                if partida['descripcion'] == descripcion
                            )

                            # Recopilar todos los materiales asociados
                            todos_materiales = []
                            for nodo in nodos_ordenados:
                                if titulo_bloque in mano_obra_por_nodo[nodo]:
                                    for partida in mano_obra_por_nodo[nodo][titulo_bloque]:
                                        if partida['descripcion'] == descripcion and partida['materiales']:
                                            todos_materiales.append(partida['materiales'])

                            # Crear la fila
                            fila = [descripcion, unidad, cantidad_total, ''] + [''] * num_nodos

                            # Agregar materiales en la columna de fecha
                            if todos_materiales:
                                fila[3] = "\n".join(todos_materiales)

                            # Para cada nodo, agregar la cantidad correspondiente
                            for i, nodo in enumerate(nodos_ordenados):
                                partida_nodo = next((
                                    partida
                                    for partida in mano_obra_por_nodo[nodo].get(titulo_bloque, [])
                                    if partida['descripcion'] == descripcion
                                ), None)

                                if partida_nodo:
                                    # Formatear contenido: cantidad + materiales
                                    contenido = f"{partida_nodo['cantidad']}"
                                    if partida_nodo['materiales']:
                                        contenido += f"\n{partida_nodo['materiales']}"

                                    fila[4 + i] = contenido

                            # Agregar la fila a la lista de filas
                            filas.append(fila)
                    
                    # Procesar otros bloques normalmente
                    elif any(titulo_bloque in mano_obra_por_nodo[nodo] for nodo in nodos_ordenados):
                        # Agregar encabezado del bloque
                        filas.append([f"BLOQUE: {titulo_bloque}", '', '', ''] + [''] * num_nodos)
                        
                        # Recopilar todas las descripciones únicas de mano de obra para este bloque
                        descripciones_unicas = set()
                        for nodo in nodos_ordenados:
                            if titulo_bloque in mano_obra_por_nodo[nodo]:
                                for partida in mano_obra_por_nodo[nodo][titulo_bloque]:
                                    descripciones_unicas.add(partida['descripcion'])
                        
                        # Para cada descripción única, crear una fila
                        for descripcion in sorted(descripciones_unicas):
                            # Encontrar la unidad (debería ser la misma para la misma descripción)
                            unidad = next((
                                partida['unidad'] 
                                for nodo in nodos_ordenados 
                                if titulo_bloque in mano_obra_por_nodo[nodo] 
                                for partida in mano_obra_por_nodo[nodo][titulo_bloque] 
                                if partida['descripcion'] == descripcion
                            ), "UND")
                            
                            # Calcular la cantidad total sumando de todos los nodos
                            cantidad_total = sum(
                                partida['cantidad']
                                for nodo in nodos_ordenados
                                if titulo_bloque in mano_obra_por_nodo[nodo]
                                for partida in mano_obra_por_nodo[nodo][titulo_bloque]
                                if partida['descripcion'] == descripcion
                            )
                            
                            # Recopilar todos los materiales asociados para mostrar en la columna de fecha
                            todos_materiales = []
                            codigos_n1 = set()
                            codigos_n2 = set()
                            brazos = set()
                            otros_materiales = set()

                            for nodo in nodos_ordenados:
                                if titulo_bloque in mano_obra_por_nodo[nodo]:
                                    for partida in mano_obra_por_nodo[nodo][titulo_bloque]:
                                        if "INSTALACION DE LUMINARIAS" in partida['descripcion'].upper() and partida['materiales']:
                                            materiales_lineas = partida['materiales'].split('\n')
                                            for linea in materiales_lineas:
                                                if "INST:" in linea:
                                                    contenido = linea.replace("INST:", "").strip()
                                                    partes = [p.strip() for p in contenido.split(',')]
                                                    for parte in partes:
                                                        if "CÓDIGO N1:" in parte:
                                                            codigo = parte.replace("CÓDIGO N1:", "").strip()
                                                            codigos_n1.add(codigo)
                                                        elif "CÓDIGO N2:" in parte:
                                                            codigo = parte.replace("CÓDIGO N2:", "").strip()
                                                            if codigo.lower() != "no":
                                                                codigos_n2.add(codigo)
                                                        elif "BRAZO" in parte:
                                                            brazos.add(parte)
                                                        else:
                                                            otros_materiales.add(parte)

                            # Crear un resumen consolidado
                            #if codigos_n1:
                            #    todos_materiales.append(f"INST: CÓDIGO N1: {', '.join(sorted(codigos_n1))}")
                            #if codigos_n2:
                            #    todos_materiales.append(f"INST: CÓDIGO N2: {', '.join(sorted(codigos_n2))}")
                            #if brazos:
                            #    todos_materiales.append(f"INST: BRAZOS: {', '.join(sorted(brazos))}")
                            #if otros_materiales:
                            #    todos_materiales.append(f"INST: OTROS: {', '.join(sorted(otros_materiales))}")
                            
                            # Crear la fila con la descripción, unidad y cantidad total
                            fila_mo = [descripcion, unidad, cantidad_total, ''] + [''] * num_nodos
                            
                            # Agregar materiales en la columna de fecha si hay alguno
                            if todos_materiales:
                                fila_mo[3] = "\n".join(todos_materiales)
                            
                            # Para cada nodo, agregar la información de mano de obra
                            for i, nodo in enumerate(nodos_ordenados):
                                # Buscar la partida para este nodo y descripción
                                partida_nodo = next((
                                    partida
                                    for partida in mano_obra_por_nodo[nodo].get(titulo_bloque, [])
                                    if partida['descripcion'] == descripcion
                                ), None)
                                
                                if partida_nodo:
                                    # IMPORTANTE: Asegurarse de que la cantidad sea un número
                                    cantidad = partida_nodo['cantidad']
                                    if isinstance(cantidad, str):
                                        try:
                                            cantidad = float(cantidad)
                                        except:
                                            cantidad = extraer_cantidad(cantidad) or 1
                                    
                                    # Formatear el contenido: cantidad + materiales
                                    contenido = f"{cantidad}"
                                    
                                    # Siempre mostrar los materiales asociados para todas las partidas
                                    if partida_nodo['materiales']:
                                        contenido += f"\n{partida_nodo['materiales']}"
                                    
                                    fila_mo[4 + i] = contenido
                                    
                                    # Imprimir información de depuración para luminarias
                                    #if "LUMINARIA" in descripcion.upper():
                                    #    print(f"DEBUG: Agregando mano de obra de luminarias en nodo {nodo}: {cantidad}")
                                    #    if partida_nodo['materiales']:
                                    #        print(f"DEBUG: Materiales: {partida_nodo['materiales']}")
                            
                            # IMPORTANTE: Agregar la fila a la lista de filas
                            filas.append(fila_mo)
                
                # Crear DataFrame a partir de los datos recopilados
                df = pd.DataFrame(filas, columns=columnas)
                df.to_excel(writer, sheet_name=f"OT_{ot}", index=False)

                # Acceder a la hoja creada
                sheet = writer.sheets[f"OT_{ot}"]
                  
                # Cargar el archivo con openpyxl para combinar celdas
                ws = writer.sheets[f"OT_{ot}"]

                # Combinar celdas para encabezados de secciones
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Comienza desde la segunda fila
                    if row[0].value in ["MATERIALES INSTALADOS", "MATERIALES RETIRADOS", "OBSERVACIONES POR NODO", 
                                       "OBSERVACIONES COMPLETAS", "MANO DE OBRA POR NODO"] or (
                                       isinstance(row[0].value, str) and row[0].value.startswith("BLOQUE: ")):
                        start_col = 1
                        end_col = len(columnas)
                        row_idx = row[0].row
                        ws.merge_cells(
                            start_row=row_idx,
                            start_column=start_col,
                            end_row=row_idx,
                            end_column=end_col
                        )
                        # Centrar el texto y aplicar formato
                        row[0].alignment = Alignment(horizontal='center', vertical='center')
                        row[0].font = Font(bold=True)
                        
                        # Aplicar color de fondo según el tipo de encabezado
                        if row[0].value == "MANO DE OBRA POR NODO":
                            row[0].fill = PatternFill("solid", fgColor="009900")  # Verde
                            row[0].font = Font(bold=True, color="FFFFFF")  # Texto blanco
                        elif isinstance(row[0].value, str) and row[0].value.startswith("BLOQUE: "):
                            row[0].fill = PatternFill("solid", fgColor="AAAAAA")  # Gris
                        else:
                            row[0].fill = PatternFill("solid", fgColor="DDDDDD")  # Gris claro
                
                # Configurar el ajuste de texto para las celdas de observaciones y mano de obra
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    row_has_multiline = False
                    max_lines = 1
                    
                    for cell in row:
                        if isinstance(cell.value, str) and '\n' in cell.value:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            row_has_multiline = True
                            lines_count = cell.value.count('\n') + 1
                            max_lines = max(max_lines, lines_count)
                    
                    # Ajustar altura de fila para acomodar el texto más largo
                    if row_has_multiline and row[0].row > 1:  # Evitar ajustar la fila de encabezados
                        # Calcular altura: 15 puntos por línea, con un mínimo de 15 y un máximo de 409
                        ws.row_dimensions[row[0].row].height = max(15, min(15 * max_lines, 409))
                
                # Asegurar que las celdas de mano de obra tengan ajuste de texto
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    if row[0].value and isinstance(row[0].value, str) and "BLOQUE:" in row[0].value:
                        # Las filas siguientes son de mano de obra hasta el próximo encabezado
                        current_row = row[0].row + 1
                        while current_row <= ws.max_row:
                            mo_row = ws[current_row]
                            if mo_row[0].value and isinstance(mo_row[0].value, str) and (
                                "BLOQUE:" in mo_row[0].value or 
                                mo_row[0].value in ["MATERIALES INSTALADOS", "MATERIALES RETIRADOS", 
                                                   "OBSERVACIONES COMPLETAS", "MANO DE OBRA POR NODO"]):
                                break
                            
                            # Configurar todas las celdas de esta fila para ajuste de texto
                            for cell in mo_row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            current_row += 1
              
            # Eliminar hoja temporal si se crearon hojas
            if sheets_created:
                writer.book.remove(writer.book[temp_sheet_name])
        
        
        # Manejo de errores
        if not sheets_created:
            writer.book.remove(writer.book[temp_sheet_name])
            df_error = pd.DataFrame({
                'Error': [
                    'Datos no válidos - Razones posibles:',
                    '1. Columnas requeridas faltantes',
                    '2. Valores "NINGUNO" o 0 en todos los registros',
                    '3. Formato de archivo incorrecto'
                ]
            })
            df_error.to_excel(writer, sheet_name='Errores', index=False)

    output.seek(0)
    return output

@app.post("/upload/")
async def subir_archivos(
    files: list[UploadFile] = File(...),
    tipo_archivo: str = Form(..., description="Tipo de archivo: modernizacion o mantenimiento")
):
    try:
        datos_combinados = defaultdict(lambda: {
            'nodos': set(),
            'codigos_n1': defaultdict(lambda: defaultdict(set)),
            'codigos_n2': defaultdict(lambda: defaultdict(set)),
            'materiales': defaultdict(lambda: defaultdict(int)),
            'materiales_retirados': defaultdict(lambda: defaultdict(int)),
            'aspectos_materiales': defaultdict(lambda: defaultdict(lambda: set())),
            'aspectos_retirados': defaultdict(lambda: defaultdict(lambda: set())),
            'fechas_sync': defaultdict(str)
        })
        
        datos_por_barrio_combinados = defaultdict(lambda: {
            'materiales_instalados': defaultdict(lambda: defaultdict(int)),
            'materiales_retirados': defaultdict(lambda: defaultdict(int)),
        })

        dfs_originales_combinados = {}  # Inicializar diccionario para DataFrames originales

        for file in files:
            try:
                if tipo_archivo == 'modernizacion':
                    # Procesar archivo y obtener datos, datos_por_barrio y DataFrames originales
                    datos, datos_por_barrio, dfs_originales = procesar_archivo_modernizacion(file)
                    # Combinar DataFrames originales
                    dfs_originales_combinados.update(dfs_originales)
                elif tipo_archivo == 'mantenimiento':
                    datos = procesar_archivo_mantenimiento(file)
                else:
                    raise ValueError("Tipo de archivo no válido")

                # Combinar datos por barrio (solo para modernización)
                if tipo_archivo == 'modernizacion':
                    for barrio, barrio_info in datos_por_barrio.items():
                        # Materiales instalados
                        for mat_key, ot_counts in barrio_info['materiales_instalados'].items():
                            for ot, cantidad in ot_counts.items():
                                datos_por_barrio_combinados[barrio]['materiales_instalados'][mat_key][ot] += cantidad
                        # Materiales retirados
                        for mat_key, ot_counts in barrio_info['materiales_retirados'].items():
                            for ot, cantidad in ot_counts.items():
                                datos_por_barrio_combinados[barrio]['materiales_retirados'][mat_key][ot] += cantidad

                # Combinar datos comunes
                for ot, info in datos.items():
                    # Actualizar nodos y fechas
                    datos_combinados[ot]['nodos'].update(info['nodos'])
                    for nodo, fecha in info.get('fechas_sync', {}).items():
                        datos_combinados[ot]['fechas_sync'][nodo] = fecha
                    
                    # Combinar materiales instalados
                    for mat_key, cantidades in info.get('materiales', {}).items():
                        for nodo, cantidad in cantidades.items():
                            datos_combinados[ot]['materiales'][mat_key][nodo] += cantidad
                    
                    # Procesar códigos y materiales retirados solo para modernización
                    if tipo_archivo == 'modernizacion':
                        # Códigos N1 y N2
                        for key, nodos_data in info.get('codigos_n1', {}).items():
                            for nodo, codigos in nodos_data.items():
                                datos_combinados[ot]['codigos_n1'][key][nodo].update(codigos)
                        for key, nodos_data in info.get('codigos_n2', {}).items():
                            for nodo, codigos in nodos_data.items():
                                datos_combinados[ot]['codigos_n2'][key][nodo].update(codigos)
                        
                        # Materiales retirados y aspectos
                        for mat_key, cantidades in info.get('materiales_retirados', {}).items():
                            for nodo, cantidad in cantidades.items():
                                datos_combinados[ot]['materiales_retirados'][mat_key][nodo] += cantidad
                        for mat_key, nodos_aspectos in info.get('aspectos_materiales', {}).items():
                            for nodo, aspectos in nodos_aspectos.items():
                                datos_combinados[ot]['aspectos_materiales'][mat_key][nodo].update(aspectos)
                        for mat_key, nodos_aspectos in info.get('aspectos_retirados', {}).items():
                            for nodo, aspectos in nodos_aspectos.items():
                                datos_combinados[ot]['aspectos_retirados'][mat_key][nodo].update(aspectos)

            except Exception as e:
                logger.error(f"Error con {file.filename}: {str(e)}")
                continue

        # Generar el Excel con todos los datos combinados
        excel_final = generar_excel(datos_combinados, datos_por_barrio_combinados, dfs_originales_combinados)
        
        return Response(
            content=excel_final.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=resultado.xlsx"}
        )

    except Exception as e:
        logger.critical(f"Error global: {str(e)}")
        raise HTTPException(500, detail=str(e))
    
      
    